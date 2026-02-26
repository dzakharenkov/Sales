import os
import json
import logging
import re
import datetime
from collections import defaultdict
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Константы в коде
SENTRY_TOKEN = os.getenv("SENTRY_TOKEN", "").strip()
ORG_SLUG = "zakharenkov"
PROJECT_ID = "4510910634196992"
BASE_URL = "https://sentry.io/api/0"
TASKS_DIR = "sentry_tasks"
EXCEL_FILE = "sentry_errors.xlsx"
BACKUP_JSON_FILE = "sentry_errors.json"
TARGET_PROJECT = "sales"
TARGET_TEAM_SLUG = "sales_zakharenkov"

if not SENTRY_TOKEN:
    raise RuntimeError("SENTRY_TOKEN is not set. Configure environment variable SENTRY_TOKEN.")

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "", filename).replace(" ", "_").replace("\n", "_").replace("\r", "_")

class SentryClient:
    def __init__(self, token, org_slug, project_id, base_url):
        self.headers = {"Authorization": f"Bearer {token}"}
        self.org_slug = org_slug
        self.project_id = project_id
        self.base_url = base_url.rstrip('/')
        
    def _make_request(self, method, url, params=None, json_data=None):
        for attempt in range(3):
            try:
                response = requests.request(
                    method,
                    url,
                    headers=self.headers,
                    params=params,
                    json=json_data,
                    timeout=10,
                )
                if response.status_code == 429: # Too many requests
                    logger.warning(f"Rate limited. Attempt {attempt + 1}/3. Waiting...")
                    import time
                    time.sleep(2)
                    continue
                response.raise_for_status()
                return response
            except requests.exceptions.RequestException as e:
                logger.error(f"Error making request to {url}: {e}")
                if attempt == 2:
                    raise
                import time
                time.sleep(1)
        return None

    def get_team_id_by_slug(self, team_slug):
        url = f"{self.base_url}/organizations/{self.org_slug}/teams/"
        response = self._make_request("GET", url, params={"query": team_slug})
        if not response:
            return None
        for team in response.json():
            if team.get("slug") == team_slug:
                return str(team.get("id"))
        return None

    def resolve_issue_with_assignee(self, issue_id, team_id):
        url = f"{self.base_url}/organizations/{self.org_slug}/issues/{issue_id}/"
        payload = {
            "status": "resolved",
            "assignedTo": f"team:{team_id}",
        }
        response = self._make_request("PUT", url, json_data=payload)
        return response is not None

    def get_all_issues(self, days=14, project_slug=None):
        issues = []
        url = f"{self.base_url}/organizations/{self.org_slug}/issues/"
        query = "is:unresolved"
        if project_slug:
            query = f"{query} project:{project_slug}"
        params = {
            "query": query,
            "statsPeriod": f"{days}d",
            "limit": 100
        }
        
        while url:
            response = self._make_request("GET", url, params=params)
            if not response:
                break
                
            data = response.json()
            issues.extend(data)
            
            # Check pagination
            links = response.links
            if 'next' in links and links['next'].get('results') == 'true':
                url = links['next']['url']
                params = None # Parameters are included in the 'next' URL
            else:
                break
                
        return issues

    def get_issue_events(self, issue_id):
        url = f"{self.base_url}/organizations/{self.org_slug}/issues/{issue_id}/events/latest/"
        response = self._make_request("GET", url)
        if response:
            return response.json()
        return None

    def _extract_stacktrace(self, event):
        stacktrace = {
            "exception_type": "Unknown",
            "exception_value": "Unknown",
            "frames": []
        }
        try:
            entries = event.get('entries', [])
            for entry in entries:
                if entry.get('type') == 'exception':
                    values = entry.get('data', {}).get('values', [])
                    if values:
                        exc = values[0]
                        stacktrace['exception_type'] = exc.get('type', 'Unknown')
                        stacktrace['exception_value'] = exc.get('value', 'Unknown')
                        
                        frames = exc.get('stacktrace', {}).get('frames', [])
                        for frame in frames:
                            stacktrace['frames'].append({
                                "filename": frame.get('filename'),
                                "line_number": frame.get('lineNo'),
                                "function": frame.get('function'),
                                "code_context": frame.get('context')
                            })
                        break
        except Exception as e:
            logger.warning(f"Error extracting stacktrace: {e}")
            
        return stacktrace

    def _extract_context(self, event):
        context = {
            "request": {},
            "user": {},
            "tags": {}
        }
        try:
            req = event.get('request', {})
            if req:
                context['request'] = {
                    "method": req.get('method'),
                    "url": req.get('url'),
                    "headers": req.get('headers', {})
                }
                
            user = event.get('user', {})
            if user:
                context['user'] = {
                    "id": user.get('id'),
                    "email": user.get('email'),
                    "ip_address": user.get('ip_address')
                }
                
            tags = event.get('tags', [])
            for tag in tags:
                if isinstance(tag, list) and len(tag) == 2:
                    context['tags'][tag[0]] = tag[1]
                elif isinstance(tag, dict) and 'key' in tag and 'value' in tag:
                    context['tags'][tag['key']] = tag['value']
        except Exception as e:
            logger.warning(f"Error extracting context: {e}")
            
        return context

    def enrich_issue(self, issue):
        event = self.get_issue_events(issue.get('id'))
        
        enriched = {
            "id": str(issue.get('id')),
            "title": issue.get('title'),
            "status": issue.get('status'),
            "level": issue.get('level'),
            "count": int(issue.get('count', '0')),
            "first_seen": issue.get('firstSeen'),
            "last_seen": issue.get('lastSeen'),
            "project": issue.get('project', {}).get('slug', 'Unknown'),
            "url": issue.get('permalink'),
            "stacktrace": {},
            "context": {}
        }
        
        if event:
            enriched["stacktrace"] = self._extract_stacktrace(event)
            enriched["context"] = self._extract_context(event)
            
        return enriched

class ExcelExporter:
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active) # Remove default sheet
        
        self.headers = [
            ("ID Ошибки", 12),
            ("Название", 40),
            ("Тип ошибки", 15),
            ("Серьёзность", 12),
            ("Кол-во возникновений", 15),
            ("Первое появление", 15),
            ("Последнее появление", 15),
            ("Файл:Строка", 30),
            ("Функция", 25),
            ("Статус", 15),
            ("Комментарий", 60),
            ("Ссылка в Sentry", 40)
        ]
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.unresolved_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.resolved_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        self.white_font = Font(color="FFFFFF")
        
    def add_project_sheet(self, project_name, issues):
        sheet_name = project_name[:31]
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.create_sheet(title=sheet_name)
            
        # Write headers
        for col_idx, (header_title, width) in enumerate(self.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_title)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            
        # Add filters
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.headers))}{len(issues) + 1}"
        
        # Write data
        for row_idx, issue in enumerate(issues, start=2):
            file_line = ""
            func = ""
            if issue['stacktrace'].get('frames'):
                last_frame = issue['stacktrace']['frames'][-1]
                if last_frame.get('filename') and last_frame.get('line_number'):
                    file_line = f"{last_frame.get('filename')}:{last_frame.get('line_number')}"
                func = last_frame.get('function', '')
                
            row_data = [
                issue['id'],
                issue['title'],
                issue['stacktrace'].get('exception_type', ''),
                issue['level'],
                issue['count'],
                issue['first_seen'][:10] if issue['first_seen'] else '',
                issue['last_seen'][:10] if issue['last_seen'] else '',
                file_line,
                func,
                "UNRESOLVED" if issue['status'] == 'unresolved' else "RESOLVED",
                issue.get('comment', ''),
                issue['url']
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 10: # Status column
                    if value == "UNRESOLVED":
                        cell.fill = self.unresolved_fill
                        cell.font = self.white_font
                    else:
                        cell.fill = self.resolved_fill
                        cell.font = self.white_font
                if col_idx == 12:
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
                    
        # Add summary rows
        start_summary_row = len(issues) + 3
        total_issues = len(issues)
        resolved = sum(1 for i in issues if i['status'] != 'unresolved')
        unresolved = total_issues - resolved
        progress = (resolved / total_issues * 100) if total_issues else 0
        
        summaries = [
            ("Всего ошибок в проекте:", total_issues),
            ("Кол-во исправленных:", resolved),
            ("Кол-во неисправленных:", unresolved),
            ("Процент готовности:", f"{progress:.1f}%")
        ]
        
        for i, (label, val) in enumerate(summaries):
            ws.cell(row=start_summary_row + i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=start_summary_row + i, column=2, value=val)

    def save(self):
        if not self.wb.sheetnames:
            self.wb.create_sheet("No Data")
        self.wb.save(self.filename)

class TaskGenerator:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.existing_ids = self._load_existing_issue_ids()

    def _load_existing_issue_ids(self):
        ids = set()
        for filename in os.listdir(self.output_dir):
            if not filename.endswith(".md"):
                continue
            match = re.match(r"^(\d+)_", filename)
            if match:
                ids.add(match.group(1))
        return ids
        
    def create_task(self, issue):
        if issue["id"] in self.existing_ids:
            return False

        safe_title = sanitize_filename(issue['title'])[:50]
        filename = f"{issue['id']}_{safe_title}.md"
        filepath = os.path.join(self.output_dir, filename)
        
        content = [
            f"# 🐛 Задача: Исправить ошибку\n",
            f"## Общая информация\n",
            f"**Проект:** {issue['project']}  ",
            f"**ID Ошибки:** {issue['id']}  ",
            f"**Ссылка в Sentry:** {issue['url']}  ",
            f"**Статус:** {'❌ НЕ ИСПРАВЛЕНО' if issue['status'] == 'unresolved' else '✅ ИСПРАВЛЕНО'}  \n",
            f"---\n",
            f"## Описание проблемы\n",
            f"**Название ошибки:**\n{issue['title']}\n"
        ]
        
        exc_type = issue['stacktrace'].get('exception_type', 'Unknown')
        exc_value = issue['stacktrace'].get('exception_value', 'Unknown')
        
        content.extend([
            f"**Тип исключения:** {exc_type}  ",
            f"**Сообщение об ошибке:** {exc_value}  \n",
            f"**Статистика:**",
            f"- Кол-во возникновений: **{issue['count']}**",
            f"- Первое появление: {issue['first_seen']}",
            f"- Последнее появление: {issue['last_seen']}",
            f"- Серьёзность: **{issue['level'].upper() if issue['level'] else 'UNKNOWN'}**\n",
            f"---\n",
            f"## Стектрейс (Stack Trace)\n"
        ])
        
        frames = issue['stacktrace'].get('frames', [])
        if frames:
            last_frame = frames[-1]
            content.extend([
                f"### Основная строка ошибки:",
                f"`{last_frame.get('filename')}:{last_frame.get('line_number')}`",
                f"в функции: `{last_frame.get('function')}`\n",
                f"### Полный стектрейс:\n"
            ])
            
            for i, frame in enumerate(reversed(frames), 1):
                content.extend([
                    f"#### Фрейм {len(frames) - i + 1}",
                    f"- **Файл:** `{frame.get('filename')}`",
                    f"- **Строка:** {frame.get('line_number')}",
                    f"- **Функция:** `{frame.get('function')}`",
                    f"- **Контекст кода:**\n",
                    f"```python"
                ])
                context = frame.get('code_context')
                if context:
                    for ctx_line in context:
                        line_content = ctx_line[1] if isinstance(ctx_line, list) and len(ctx_line)>1 else str(ctx_line)
                        content.append(line_content)
                content.append(f"```\n")
        else:
            content.append("Стектрейс недоступен.\n")
            
        content.append("---\n")
        content.append("## Контекстная информация\n")
        
        ctx = issue['context']
        req = ctx.get('request', {})
        if req:
            content.extend([
                f"### HTTP Запрос",
                f"**Метод:** {req.get('method')}",
                f"**URL:** {req.get('url')}"
            ])
            headers = req.get('headers', {})
            if headers:
                content.append("**Headers:**")
                content.append("```json")
                content.append(json.dumps(headers, indent=2))
                content.append("```\n")
                
        user = ctx.get('user', {})
        if user:
            content.extend([
                f"### Информация о пользователе",
                f"- **User ID:** {user.get('id', 'N/A')}",
                f"- **Email:** {user.get('email', 'N/A')}",
                f"- **IP Address:** {user.get('ip_address', 'N/A')}\n"
            ])
            
        tags = ctx.get('tags', {})
        if tags:
            content.append("### Теги")
            content.append("```json")
            content.append(json.dumps(tags, indent=2))
            content.append("```\n")
            
        content.extend([
            f"---\n",
            f"## Инструкции для исправления\n",
            f"**Воспроизведение ошибки:**",
            f"1. Перейти по ссылке в Sentry для просмотра деталей",
            f"2. Понять сценарий, при котором происходит ошибка",
            f"3. Попытаться воспроизвести локально\n",
            f"**Анализ кода:**",
            f"1. Открыть файл на строке, указанной в стектрейсе",
            f"2. Проанализировать логику и найти причину",
            f"3. Проверить граничные случаи и исключения\n",
            f"**Исправление:**",
            f"1. Написать патч/исправление",
            f"2. Добавить проверку входных данных, если необходимо",
            f"3. Убедиться, что обработка исключений корректна\n",
            f"**Тестирование:**",
            f"1. Написать юнит-тест, покрывающий этот сценарий",
            f"2. Провести ручное тестирование",
            f"3. Убедиться, что ошибка больше не возникает\n",
            f"**После исправления:**",
            f"1. Создать commit с описанием",
            f"2. Развернуть в production",
            f"3. Отметить статус в Excel как ✅ ИСПРАВЛЕНО\n",
            f"---\n",
            f"## Статус выполнения\n",
            f"- [ ] Ошибка воспроизведена",
            f"- [ ] Найдена причина",
            f"- [ ] Написано исправление",
            f"- [ ] Написаны тесты",
            f"- [ ] Code review пройден",
            f"- [ ] Развёрнуто в production",
            f"- [ ] Проверено в Sentry (ошибка исчезла)\n",
            f"**Статус:** {'❌ НЕ НАЧИНАЛОСЬ' if issue['status'] == 'unresolved' else '✅ ЗАВЕРШЕНО'}\n\n",
            f"Комментарий решения: \n\n",
            f"Создано: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ])
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("\n".join(content))
        self.existing_ids.add(issue["id"])
        return True
            
def load_previous_issues():
    if not os.path.exists(BACKUP_JSON_FILE):
        return []

    try:
        with open(BACKUP_JSON_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as exc:
        logger.warning(f"Cannot read previous backup '{BACKUP_JSON_FILE}': {exc}")
        return []

    return data.get("errors", [])


def merge_issues(previous_issues, current_issues, project_slug):
    merged = {}

    for issue in previous_issues:
        if issue.get("project") != project_slug:
            continue
        issue_copy = dict(issue)
        issue_copy["status"] = "resolved"
        merged[str(issue_copy["id"])] = issue_copy

    current_ids = set()
    for issue in current_issues:
        issue_id = str(issue["id"])
        issue["status"] = "unresolved"
        merged[issue_id] = issue
        current_ids.add(issue_id)

    # Keep historical issues in the table; mark missing current issues as resolved.
    for issue_id, issue in merged.items():
        if issue_id not in current_ids:
            issue["status"] = "resolved"

    return sorted(
        merged.values(),
        key=lambda item: (item.get("last_seen") or "", item.get("first_seen") or ""),
        reverse=True,
    )


def get_resolved_task_ids(tasks_dir):
    resolved_ids = set()
    if not os.path.isdir(tasks_dir):
        return resolved_ids

    for filename in os.listdir(tasks_dir):
        if not filename.endswith(".md"):
            continue
        if "_RESOLVED" not in filename:
            continue
        match = re.match(r"^(\d+)_", filename)
        if match:
            resolved_ids.add(match.group(1))
    return resolved_ids


def load_task_comments(tasks_dir):
    comments = {}
    if not os.path.isdir(tasks_dir):
        return comments

    for filename in os.listdir(tasks_dir):
        if not filename.endswith(".md"):
            continue

        match = re.match(r"^(\d+)_", filename)
        if not match:
            continue
        issue_id = match.group(1)
        file_path = os.path.join(tasks_dir, filename)

        comment = ""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                for raw_line in f:
                    line = raw_line.strip()
                    if line.startswith("Комментарий решения:") or line.startswith("**Комментарий решения:**"):
                        comment = line.split(":", 1)[1].strip()
                        break
        except OSError:
            continue

        if not comment and "_RESOLVED" in filename:
            comment = "Исправлено в проекте, задача закрыта и синхронизирована в Sentry."

        if comment:
            comments[issue_id] = comment

    return comments


def create_backup(issues, projects_stats):
    backup_data = {
        "metadata": {
            "organization": ORG_SLUG,
            "project_filter": TARGET_PROJECT,
            "total_issues": len(issues),
            "fetched_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "period": "14d unresolved + historical merged",
            "by_project": projects_stats
        },
        "errors": issues
    }
    with open(BACKUP_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2)

def main():
    print("🚀 Запуск обработки ошибок Sentry...")

    client = SentryClient(SENTRY_TOKEN, ORG_SLUG, PROJECT_ID, BASE_URL)

    resolved_task_ids = get_resolved_task_ids(TASKS_DIR)
    task_comments = load_task_comments(TASKS_DIR)
    team_id = client.get_team_id_by_slug(TARGET_TEAM_SLUG)
    synced_to_portal = 0
    if resolved_task_ids and team_id:
        print("🔧 Синхронизирую локальные RESOLVED-задачи с Sentry...")
        for issue_id in sorted(resolved_task_ids):
            if client.resolve_issue_with_assignee(issue_id, team_id):
                synced_to_portal += 1
        print(f"✅ В Sentry обновлено задач: {synced_to_portal}")
    elif resolved_task_ids and not team_id:
        print(f"⚠️ Команда '{TARGET_TEAM_SLUG}' не найдена: assignee не будет обновлен")

    print("🔍 Получаю список ошибок из Sentry...")
    raw_issues = client.get_all_issues(days=14, project_slug=TARGET_PROJECT)
    if not isinstance(raw_issues, list):
        print("❌ Ошибка при получении данных от Sentry")
        return
        
    print(f"✅ Получено {len(raw_issues)} ошибок")
    
    enriched_issues = []
    for i, raw_issue in enumerate(raw_issues, 1):
        print(f"[{i}/{len(raw_issues)}] Обработка ошибки...")
        logger.info(f"Обогащаю ошибку {raw_issue.get('id')}: {raw_issue.get('title')}")
        enriched = client.enrich_issue(raw_issue)
        enriched_issues.append(enriched)
    
    print(f"✅ Обработано {len(enriched_issues)} ошибок (sales unresolved)")

    previous_issues = load_previous_issues()
    previous_sales_ids = {
        str(issue.get("id"))
        for issue in previous_issues
        if issue.get("project") == TARGET_PROJECT
    }
    merged_issues = merge_issues(previous_issues, enriched_issues, TARGET_PROJECT)
    current_sales_ids = {str(issue["id"]) for issue in enriched_issues}
    new_issue_ids = current_sales_ids - previous_sales_ids

    resolved_from_tasks = 0
    for issue in merged_issues:
        issue_id = str(issue["id"])
        issue["comment"] = task_comments.get(issue_id, "")
        if str(issue["id"]) in resolved_task_ids:
            issue["status"] = "resolved"
            resolved_from_tasks += 1

    print(f"📚 История Sales: {len(merged_issues)} (новых: {len(new_issue_ids)})")
    print(f"✅ Локально помечено RESOLVED по задачам: {resolved_from_tasks}")
    
    print("📊 Группировка данных и создание Excel файла...")
    issues_by_project = defaultdict(list)
    for issue in merged_issues:
        issues_by_project[issue['project']].append(issue)
        
    projects_stats = {}
    exporter = ExcelExporter(EXCEL_FILE)
    for project_name, prj_issues in issues_by_project.items():
        projects_stats[project_name] = len(prj_issues)
        exporter.add_project_sheet(project_name, prj_issues)
        print(f"✅ Добавлен лист '{project_name}' с {len(prj_issues)} ошибками")
        
    exporter.save()
    print(f"💾 Excel файл сохранён: {EXCEL_FILE}")
    
    print("📝 Создаю резервную JSON копию...")
    create_backup(merged_issues, projects_stats)
    print(f"💾 JSON файл сохранён: {BACKUP_JSON_FILE}")
    
    print("📝 Создаю Markdown задачи...")
    generator = TaskGenerator(TASKS_DIR)
    created_tasks = 0
    for issue in merged_issues:
        if issue["status"] != "unresolved":
            continue
        if issue["id"] not in new_issue_ids:
            continue
        if generator.create_task(issue):
            created_tasks += 1

    print(f"✅ Созданы новые задачи в папке {TASKS_DIR}/: {created_tasks}")
    
    print("\n╔" + "═" * 56 + "╗")
    print("║           ✅ ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО             ║")
    print("╚" + "═" * 56 + "╝\n")
    
    print("📊 Статистика:")
    print(f"  • Всего ошибок в базе Sales: {len(merged_issues)}")
    print(f"  • Проектов: {len(issues_by_project)}")
    print(f"  • Excel файл: {EXCEL_FILE}")
    print(f"  • JSON бэкап: {BACKUP_JSON_FILE}")
    print(f"  • Папка с задачами: {TASKS_DIR}/")

if __name__ == "__main__":
    main()

"""
ÐžÐ¿ÐµÑ€Ð°Ñ†Ð¸Ð¸: Ð½Ð¾Ð²Ð°Ñ ÑÑ‚Ñ€ÑƒÐºÑ‚ÑƒÑ€Ð° (operation_number, type_code, warehouse_from/to, status Ð¸ Ð´Ñ€.).
"""
import io
from datetime import date, datetime, timezone
from uuid import UUID
from src.api.v1.schemas.common import EntityModel
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from pydantic import BaseModel, field_validator
from sqlalchemy import select, text, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy.exc import IntegrityError

from src.database.connection import get_db_session
from src.database.models import Operation, Customer, Product, Order, OperationConfig, OperationType, Batch, Item, Status
from src.core.deps import get_current_user, require_admin
from src.core.pagination import PaginatedResponse, PaginationParams
from src.database.models import User
from src.api.v1.services.operation_service import OperationService
from src.api.v1.services.translation_service import TranslationService
import json
import logging

router = APIRouter()
ALLOWED_USER_ROLES = {"admin", "expeditor", "agent", "stockman", "paymaster"}
OUTFLOW_OPERATION_TYPES = {
    "allocation",
    "delivery",
    "promotional_sample",
    "write_off",
    "damage",
    "transfer",
}


def _parse_delivery_date_input(raw: str) -> date:
    """Parse delivery date from yyyy-mm-dd or dd.mm.yyyy."""
    value = (raw or "").strip()
    if not value:
        raise HTTPException(status_code=400, detail="ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð° Ð´Ð°Ñ‚Ð° Ð¿Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸.")
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise HTTPException(
        status_code=400,
        detail="ÐÐµÐ²ÐµÑ€Ð½Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚ Ð´Ð°Ñ‚Ñ‹ Ð¿Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸. Ð˜ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐ¹Ñ‚Ðµ Ð´Ð´.Ð¼Ð¼.Ð³Ð³Ð³Ð³ Ð¸Ð»Ð¸ yyyy-mm-dd.",
    )


async def _ensure_user_can_create_operation_type(
    session: AsyncSession,
    user: User,
    operation_type: str,
) -> str:
    col_result = await session.execute(
        text(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'Sales'
              AND table_name = 'operation_types'
              AND column_name = 'executor_role'
            LIMIT 1
            """
        )
    )
    has_executor_role = col_result.scalar_one_or_none() is not None

    if has_executor_role:
        op_type_result = await session.execute(
            text('SELECT code, COALESCE(executor_role, \'\') FROM "Sales".operation_types WHERE code = :code'),
            {"code": operation_type},
        )
        op_type_row = op_type_result.fetchone()
    else:
        op_type_result = await session.execute(
            text('SELECT code FROM "Sales".operation_types WHERE code = :code'),
            {"code": operation_type},
        )
        row = op_type_result.fetchone()
        legacy_roles = {
            "payment_receipt_from_customer": "expeditor",
            "cash_receipt": "paymaster",
            "warehouse_receipt": "stockman",
            "allocation": "stockman",
            "transfer": "stockman",
            "write_off": "stockman",
            "delivery": "expeditor",
            "promotional_sample": "expeditor",
        }
        op_type_row = (row[0], legacy_roles.get(operation_type, "admin")) if row else None

    if not op_type_row:
        raise HTTPException(status_code=404, detail=f"Operation type '{operation_type}' not found")

    current_role = (user.role or "").strip().lower()
    required_role = (op_type_row[1] or "").strip().lower()
    if current_role == "admin":
        return required_role
    if not required_role:
        raise HTTPException(
            status_code=403,
            detail="Executor role is not configured for this operation type. Contact administrator.",
        )
    if current_role != required_role:
        raise HTTPException(
            status_code=403,
            detail=f"Only role '{required_role}' (or admin) can create operation type '{operation_type}'.",
        )
    return required_role


async def _ensure_sufficient_batch_stock(
    session: AsyncSession,
    operation_type: str,
    warehouse_from: str | None,
    product_code: str | None,
    batch_code: str | None,
    qty: int | float | None,
) -> None:
    if operation_type not in OUTFLOW_OPERATION_TYPES:
        return
    if not warehouse_from or not product_code or not batch_code:
        return
    requested = int(qty or 0)
    if requested <= 0:
        return

    stock_result = await session.execute(
        text(
            '''
            SELECT COALESCE(total_qty, 0)::int
            FROM "Sales".v_warehouse_stock
            WHERE warehouse_code = :wh
              AND product_code = :pc
              AND batch_code = :bc
            '''
        ),
        {"wh": warehouse_from, "pc": product_code, "bc": batch_code},
    )
    stock_row = stock_result.fetchone()
    available = int(stock_row[0]) if stock_row and stock_row[0] is not None else 0
    if available < requested:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Недостаточно остатков для операции '{operation_type}': "
                f"склад={warehouse_from}, товар={product_code}, партия={batch_code}, "
                f"доступно {available}, запрошено {requested}."
            ),
        )


@router.get("/operations/allocation/suggest-by-delivery-date", response_model=EntityModel | list[EntityModel])
async def suggest_allocation_items_by_delivery_date(
    warehouse_from: str = Query(..., description="ÐšÐ¾Ð´ ÑÐºÐ»Ð°Ð´Ð° Ð¾Ñ‚"),
    expeditor_login: str = Query(..., description="Ð›Ð¾Ð³Ð¸Ð½ ÑÐºÑÐ¿ÐµÐ´Ð¸Ñ‚Ð¾Ñ€Ð°"),
    delivery_date: str = Query(..., description="Ð”Ð°Ñ‚Ð° Ð¿Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ (yyyy-mm-dd Ð¸Ð»Ð¸ dd.mm.yyyy)"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """ÐÐ²Ñ‚Ð¾Ð¿Ð¾Ð´Ð±Ð¾Ñ€ Ð¿Ð¾Ð·Ð¸Ñ†Ð¸Ð¹ Ð´Ð»Ñ allocation Ð½Ð° Ð¾ÑÐ½Ð¾Ð²Ðµ Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð² Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐµ Ð¿Ð¾ Ð´Ð°Ñ‚Ðµ Ð¿Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸."""
    delivery_dt = _parse_delivery_date_input(delivery_date)

    # 1) ÐÐ°Ñ…Ð¾Ð´Ð¸Ð¼ Ð’Ð¡Ð• Ð°ÐºÑ‚Ð¸Ð²Ð½Ñ‹Ðµ (Ð½Ðµ Ð·Ð°ÐºÑ€Ñ‹Ñ‚Ñ‹Ðµ) Ð·Ð°ÐºÐ°Ð·Ñ‹ ÑÐºÑÐ¿ÐµÐ´Ð¸Ñ‚Ð¾Ñ€Ð° Ð½Ð° Ð²Ñ‹Ð±Ñ€Ð°Ð½Ð½ÑƒÑŽ Ð´Ð°Ñ‚Ñƒ.
    status_code_l = func.lower(func.coalesce(Order.status_code, ""))
    status_name_l = func.lower(func.coalesce(Status.name, ""))
    matched_orders_result = await session.execute(
        select(Order.order_no)
        .select_from(Order)
        .join(Customer, Customer.id == Order.customer_id)
        .outerjoin(Status, Status.code == Order.status_code)
        .where(Customer.login_expeditor == expeditor_login.strip())
        .where(func.date(Order.scheduled_delivery_at) == delivery_dt)
        .where(
            ~or_(
                status_code_l.in_(["completed", "cancelled", "canceled", "3", "4"]),
                status_name_l.like("%Ð¾Ñ‚Ð¼ÐµÐ½%"),
                status_name_l.like("%Ð´Ð¾ÑÑ‚Ð°Ð²Ð»ÐµÐ½%"),
                status_name_l.like("%cancel%"),
                status_name_l.like("%completed%"),
                status_name_l.like("%closed%"),
            )
        )
    )
    matched_order_ids = [int(r[0]) for r in matched_orders_result.fetchall() if r and r[0] is not None]
    if not matched_order_ids:
        return {
            "success": True,
            "delivery_date": delivery_dt.isoformat(),
            "expeditor_login": expeditor_login,
            "warehouse_from": warehouse_from,
            "no_orders": True,
            "matched_orders_count": 0,
            "matched_order_ids": [],
            "message": f"ÐÐ° Ð²Ñ‹Ð±Ñ€Ð°Ð½Ð½ÑƒÑŽ Ð´Ð°Ñ‚Ñƒ Ð½ÐµÑ‚ Ð°ÐºÑ‚Ð¸Ð²Ð½Ñ‹Ñ… Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð´Ð»Ñ Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ ÑÐºÑÐ¿ÐµÐ´Ð¸Ñ‚Ð¾Ñ€Ð¾Ð¼ {expeditor_login}.",
            "items": [],
            "warnings": [],
        }

    # 2) Ð¡ÑƒÐ¼Ð¼Ð¸Ñ€ÑƒÐµÐ¼ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð° Ð¿Ð¾ Ñ‚Ð¾Ð²Ð°Ñ€Ð°Ð¼ Ð¿Ð¾ Ð’Ð¡Ð•Ðœ Ð½Ð°Ð¹Ð´ÐµÐ½Ð½Ñ‹Ð¼ Ð·Ð°ÐºÐ°Ð·Ð°Ð¼.
    grouped_items_result = await session.execute(
        select(
            Item.product_code,
            func.sum(func.coalesce(Item.quantity, 0)).label("required_qty"),
        )
        .select_from(Item)
        .where(Item.order_id.in_(matched_order_ids))
        .group_by(Item.product_code)
    )
    filtered_required: dict[str, int] = {}
    for product_code, required_qty in grouped_items_result.fetchall():
        if not product_code:
            continue
        filtered_required[product_code] = int(required_qty or 0)

    product_codes = list(filtered_required.keys())

    # 2) ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð¾ÑÑ‚Ð°Ñ‚ÐºÐ¸ Ð¿Ð¾ Ð¿Ð°Ñ€Ñ‚Ð¸ÑÐ¼ ÑÐ¾ ÑÐºÐ»Ð°Ð´Ð° Ð¾Ñ‚.
    stock_result = await session.execute(
        text(
            '''
            SELECT warehouse_code, product_code, batch_code, batch_id, total_qty
            FROM "Sales".v_warehouse_stock
            WHERE warehouse_code = :warehouse_from
              AND total_qty > 0
              AND product_code = ANY(:product_codes)
            ORDER BY product_code, batch_code
            '''
        ),
        {"warehouse_from": warehouse_from, "product_codes": product_codes},
    )
    stock_rows = stock_result.fetchall()

    by_product_stock: dict[str, list[dict]] = {pc: [] for pc in product_codes}
    batch_ids = set()
    for wh_code, product_code, batch_code, batch_id, total_qty in stock_rows:
        if not product_code:
            continue
        if batch_id:
            batch_ids.add(batch_id)
        by_product_stock.setdefault(product_code, []).append(
            {
                "warehouse_code": wh_code,
                "product_code": product_code,
                "batch_code": batch_code,
                "batch_id": batch_id,
                "available_qty": int(total_qty or 0),
            }
        )

    # 3) ÐŸÐ¾Ð´Ñ‚ÑÐ³Ð¸Ð²Ð°ÐµÐ¼ ÑÐ¿Ñ€Ð°Ð²Ð¾Ñ‡Ð½Ñ‹Ðµ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð¾ ÑÑ€Ð¾ÐºÐ°Ñ… Ð¸ Ñ‚Ð¾Ð²Ð°Ñ€Ð°Ñ….
    expiry_by_batch: dict = {}
    if batch_ids:
        batch_result = await session.execute(
            select(Batch.id, Batch.expiry_date).where(Batch.id.in_(batch_ids))
        )
        for b_id, expiry_date in batch_result.fetchall():
            expiry_by_batch[b_id] = expiry_date

    product_info: dict[str, dict] = {}
    product_result = await session.execute(
        select(Product.code, Product.name, Product.price, Product.weight_g).where(Product.code.in_(product_codes))
    )
    for code, name, price, weight_g in product_result.fetchall():
        product_info[code] = {
            "product_name": name or code,
            "unit_price": float(price) if price is not None else 0.0,
            "weight_g": int(weight_g) if weight_g is not None else 0,
        }

    # 4) FEFO-Ð°Ð»Ð³Ð¾Ñ€Ð¸Ñ‚Ð¼: Ñ€Ð°ÑÐ¿Ñ€ÐµÐ´ÐµÐ»ÑÐµÐ¼ Ñ‚Ñ€ÐµÐ±ÑƒÐµÐ¼Ð¾Ðµ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð¿Ð¾ Ð´Ð¾ÑÑ‚ÑƒÐ¿Ð½Ñ‹Ð¼ Ð¿Ð°Ñ€Ñ‚Ð¸ÑÐ¼.
    today = date.today()
    items: list[dict] = []
    warnings: list[str] = []

    for product_code, required_qty in filtered_required.items():
        stock_list = by_product_stock.get(product_code, [])
        for row in stock_list:
            exp_date = expiry_by_batch.get(row.get("batch_id"))
            row["expiry_date"] = exp_date
            row["days_until_expiry"] = (exp_date - today).days if exp_date else None
        stock_list.sort(
            key=lambda r: (r["days_until_expiry"] if r["days_until_expiry"] is not None else 10**9)
        )

        remaining = int(required_qty or 0)
        allocated_total = 0
        available_total = sum(int(r.get("available_qty") or 0) for r in stock_list)
        product_meta = product_info.get(
            product_code,
            {"product_name": product_code, "unit_price": 0.0, "weight_g": 0},
        )

        allocations: list[dict] = []
        for st in stock_list:
            if remaining <= 0:
                break
            available = int(st.get("available_qty") or 0)
            if available <= 0:
                continue
            take_qty = min(remaining, available)
            remaining -= take_qty
            allocated_total += take_qty
            allocations.append(
                {
                    "product_code": product_code,
                    "product_name": product_meta["product_name"],
                    "batch_code": st.get("batch_code"),
                    "expiry_date": st.get("expiry_date").isoformat() if st.get("expiry_date") else None,
                    "days_until_expiry": st.get("days_until_expiry"),
                    "available_qty": available,
                    "required_qty": int(required_qty or 0),
                    "allocated_qty": take_qty,
                    "unit_price": product_meta["unit_price"],
                    "weight_g": product_meta["weight_g"],
                }
            )

        shortage_qty = max(0, int(required_qty or 0) - allocated_total)
        if shortage_qty > 0:
            warnings.append(
                f"Ð¢Ð¾Ð²Ð°Ñ€ '{product_meta['product_name']}' Ð½ÐµÐ´Ð¾ÑÑ‚Ð°Ñ‚Ð¾Ñ‡Ð½Ð¾ Ð½Ð° ÑÐºÐ»Ð°Ð´Ðµ: "
                f"Ñ‚Ñ€ÐµÐ±ÑƒÐµÑ‚ÑÑ {int(required_qty or 0)} ÑˆÑ‚., Ð´Ð¾ÑÑ‚ÑƒÐ¿Ð½Ð¾ {available_total} ÑˆÑ‚., Ð·Ð°Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¾ {allocated_total} ÑˆÑ‚."
            )

        items.append(
            {
                "product_code": product_code,
                "product_name": product_meta["product_name"],
                "required_qty": int(required_qty or 0),
                "available_qty_total": available_total,
                "allocated_qty_total": allocated_total,
                "shortage_qty": shortage_qty,
                "allocations": allocations,
            }
        )

    return {
        "success": True,
        "delivery_date": delivery_dt.isoformat(),
        "expeditor_login": expeditor_login,
        "warehouse_from": warehouse_from,
        "no_orders": False,
        "matched_orders_count": len(matched_order_ids),
        "matched_order_ids": matched_order_ids,
        "items": items,
        "warnings": warnings,
    }


@router.get("/operation-types", response_model=EntityModel | list[EntityModel])
async def list_operation_types(
    language: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Ð¢Ð¸Ð¿Ñ‹ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹: Ð¿Ñ€Ð¸Ñ…Ð¾Ð´, Ñ€Ð°ÑÑ…Ð¾Ð´, Ð¿Ñ€Ð¾Ð´Ð°Ð¶Ð°, Ð²Ð¾Ð·Ð²Ñ€Ð°Ñ‚ Ð¸ Ñ‚.Ð´. (code PK) + Ð°ÐºÑ‚Ð¸Ð²Ð½Ð¾ÑÑ‚ÑŒ Ð¸Ð· operation_config.
    active=True Ñ‚Ð¾Ð»ÑŒÐºÐ¾ ÐµÑÐ»Ð¸ ÐµÑÑ‚ÑŒ operation_config Ð¸ oc.active=TRUE. has_config=True ÐµÑÐ»Ð¸ ÐºÐ¾Ð½Ñ„Ð¸Ð³ ÐµÑÑ‚ÑŒ (Ð´Ð»Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸)."""
    col_result = await session.execute(
        text(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'Sales'
              AND table_name = 'operation_types'
              AND column_name = 'executor_role'
            LIMIT 1
            """
        )
    )
    has_executor_role = col_result.scalar_one_or_none() is not None

    result = await session.execute(
        text(
            '''
            SELECT ot.code, ot.name, ot.description,
                   '''
            + ("ot.executor_role" if has_executor_role else "NULL::text")
            + ''',
                   (oc.operation_type_code IS NOT NULL AND COALESCE(oc.active, TRUE) = TRUE) AS active,
                   (oc.operation_type_code IS NOT NULL) AS has_config
            FROM "Sales".operation_types ot
            LEFT JOIN "Sales".operation_config oc
              ON oc.operation_type_code = ot.code
            ORDER BY ot.code
            '''
        )
    )
    rows = result.fetchall()
    translation_service = TranslationService(session)
    keys = [f"operation_type.{r[0]}" for r in rows if r[0]]
    translated = await translation_service.resolve_many(keys, language)
    return [
        {
            "code": r[0],
            "name": translated.get(f"operation_type.{r[0]}", r[1]),
            "description": r[2],
            "role": r[3],
            "active": bool(r[4]),
            "has_config": bool(r[5]),
        }
        for r in rows
    ]


async def _has_executor_role_column(session: AsyncSession) -> bool:
    col_result = await session.execute(
        text(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'Sales'
              AND table_name = 'operation_types'
              AND column_name = 'executor_role'
            LIMIT 1
            """
        )
    )
    return col_result.scalar_one_or_none() is not None


class OperationTypeCreate(BaseModel):
    code: str
    name: str
    description: str | None = None
    role: str

    @field_validator("role")
    @classmethod
    def validate_role(cls, value: str) -> str:
        normalized = (value or "").strip().lower()
        if normalized not in ALLOWED_USER_ROLES:
            raise ValueError("role must be one of: admin, expeditor, agent, stockman, paymaster")
        return normalized
    # ÐÐºÑ‚Ð¸Ð²Ð½Ð¾ÑÑ‚ÑŒ Ð±ÐµÑ€Ñ‘Ñ‚ÑÑ Ð¸Ð· operation_config (active), Ð·Ð´ÐµÑÑŒ Ð½Ðµ ÑƒÐ¿Ñ€Ð°Ð²Ð»ÑÐµÐ¼


class OperationTypeUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    role: str | None = None
    active: bool | None = None

    @field_validator("role")
    @classmethod
    def validate_role(cls, value: str | None) -> str | None:
        if value is None:
            return None
        normalized = value.strip().lower()
        if normalized not in ALLOWED_USER_ROLES:
            raise ValueError("role must be one of: admin, expeditor, agent, stockman, paymaster")
        return normalized


@router.post("/operation-types", response_model=EntityModel | list[EntityModel])
async def create_operation_type(
    body: OperationTypeCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Ð”Ð¾Ð±Ð°Ð²Ð¸Ñ‚ÑŒ Ñ‚Ð¸Ð¿ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸. Ð¢Ð¾Ð»ÑŒÐºÐ¾ admin. Ð¢Ð°Ð±Ð»Ð¸Ñ†Ð° Sales.operation_types (code PK)."""
    if await _has_executor_role_column(session):
        await session.execute(
            text('INSERT INTO "Sales".operation_types (code, name, description, executor_role) VALUES (:code, :name, :desc, :role)'),
            {"code": body.code, "name": body.name, "desc": body.description, "role": body.role},
        )
    else:
        await session.execute(
            text('INSERT INTO "Sales".operation_types (code, name, description) VALUES (:code, :name, :desc)'),
            {"code": body.code, "name": body.name, "desc": body.description},
        )
    await session.commit()
    return {"code": body.code, "name": body.name, "role": body.role, "message": "created"}


@router.put("/operation-types/{code}")
async def update_operation_type(
    code: str,
    body: OperationTypeUpdate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Ð˜Ð·Ð¼ÐµÐ½Ð¸Ñ‚ÑŒ Ñ‚Ð¸Ð¿ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸. Ð¢Ð¾Ð»ÑŒÐºÐ¾ admin. ÐÐºÑ‚Ð¸Ð²Ð½Ð¾ÑÑ‚ÑŒ Ð±ÐµÑ€Ñ‘Ñ‚ÑÑ Ð¸Ð· operation_config.active."""
    if body.name is not None:
        await session.execute(
            text('UPDATE "Sales".operation_types SET name = :name WHERE code = :code'),
            {"name": body.name, "code": code},
        )
    if body.description is not None:
        await session.execute(
            text('UPDATE "Sales".operation_types SET description = :desc WHERE code = :code'),
            {"desc": body.description, "code": code},
        )
    if body.role is not None and await _has_executor_role_column(session):
        await session.execute(
            text('UPDATE "Sales".operation_types SET executor_role = :role WHERE code = :code'),
            {"role": body.role, "code": code},
        )
    if body.active is not None:
        # ÐžÐ±Ð½Ð¾Ð²Ð»ÑÐµÐ¼ Ð°ÐºÑ‚Ð¸Ð²Ð½Ð¾ÑÑ‚ÑŒ Ð² Ñ‚Ð°Ð±Ð»Ð¸Ñ†Ðµ ÐºÐ¾Ð½Ñ„Ð¸Ð³ÑƒÑ€Ð°Ñ†Ð¸Ð¸
        await session.execute(
            text('UPDATE "Sales".operation_config SET active = :act WHERE operation_type_code = :code'),
            {"act": body.active, "code": code},
        )
    await session.commit()
    return {"code": code, "message": "updated"}


@router.delete("/operation-types/{code}")
async def delete_operation_type(
    code: str,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Delete operation type. Admin only."""
    exists_result = await session.execute(
        text('SELECT 1 FROM "Sales".operation_types WHERE code = :code'),
        {"code": code},
    )
    if exists_result.fetchone() is None:
        raise HTTPException(status_code=404, detail="Тип операции не найден")

    usage_result = await session.execute(
        text('SELECT COUNT(*) FROM "Sales".operations WHERE type_code = :code'),
        {"code": code},
    )
    usage_count = int(usage_result.scalar() or 0)
    if usage_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Нельзя удалить тип операции '{code}': есть связанные операции ({usage_count}).",
        )

    try:
        await session.execute(
            text('DELETE FROM "Sales".operation_config WHERE operation_type_code = :code'),
            {"code": code},
        )
        result = await session.execute(
            text('DELETE FROM "Sales".operation_types WHERE code = :code'),
            {"code": code},
        )
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Нельзя удалить тип операции '{code}': есть связанные данные.",
        )

    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Тип операции не найден")
    return {"code": code, "message": "deleted"}


@router.get("/operations", response_model=PaginatedResponse[EntityModel])
async def list_operations(
    type_code: str | None = Query(None),
    customer_id: int | None = Query(None),
    product_code: str | None = Query(None),
    status: str | None = Query(None),
    from_date: date | None = Query(None),
    to_date: date | None = Query(None),
    created_by: str | None = Query(None),
    pagination: PaginationParams = Depends(),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Ð¡Ð¿Ð¸ÑÐ¾Ðº Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹ Ñ Ñ„Ð¸Ð»ÑŒÑ‚Ñ€Ð°Ð¼Ð¸."""
    data, total = await OperationService(session).list_operations(
        pagination,
        type_code=type_code,
        customer_id=customer_id,
        product_code=product_code,
        status=status,
        from_date=from_date,
        to_date=to_date,
        created_by=created_by,
    )
    return PaginatedResponse.create(data=data, total=total, pagination=pagination)
@router.get("/operations/export", response_model=None)
async def export_operations_excel(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Ð’Ñ‹Ð³Ñ€ÑƒÐ·ÐºÐ° Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹ Ð² Excel. Ð¢Ð¾Ð»ÑŒÐºÐ¾ admin."""
    q = select(Operation).order_by(Operation.operation_date.desc().nulls_last(), Operation.created_at.desc().nulls_last())
    result = await session.execute(q)
    rows = result.scalars().all()[:50000]
    types_name_map = {}
    t_result = await session.execute(text('SELECT code, name FROM "Sales".operation_types'))
    for row in t_result.fetchall():
        types_name_map[row[0]] = row[1] or row[0]
    customer_ids = {r.customer_id for r in rows if r.customer_id is not None}
    customer_names = {}
    if customer_ids:
        c_result = await session.execute(select(Customer.id, Customer.name_client, Customer.firm_name).where(Customer.id.in_(customer_ids)))
        for r in c_result.all():
            customer_names[r[0]] = (r[1] or r[2] or "")
    wb = Workbook()
    ws = wb.active
    ws.title = "ÐžÐ¿ÐµÑ€Ð°Ñ†Ð¸Ð¸"
    for col, name in enumerate(OPERATIONS_EXPORT_HEADERS_RU, start=1):
        ws.cell(row=1, column=col, value=name)
    for row_idx, o in enumerate(rows, start=2):
        type_name = types_name_map.get(o.type_code) or o.type_code or ""
        cust_name = customer_names.get(o.customer_id) or ""
        row_data = [
            o.operation_number or "",
            o.operation_date.isoformat() if o.operation_date else "",
            type_name,
            o.status or "",
            o.warehouse_from or "",
            o.warehouse_to or "",
            o.product_code or "",
            o.quantity or "",
            float(o.amount) if o.amount is not None else "",
            cust_name,
            o.order_id or "",
            o.created_by or "",
            o.comment or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="operations.xlsx"'},
    )


@router.get("/operations/{operation_id}", response_model=EntityModel | list[EntityModel])
async def get_operation(
    operation_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """ÐžÐ´Ð½Ð° Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ñ Ð¿Ð¾ id."""
    result = await session.execute(select(Operation).where(Operation.id == operation_id))
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Operation not found")
    type_name = None
    if op.type_code:
        t = await session.execute(
            text('SELECT name FROM "Sales".operation_types WHERE code = :code'),
            {"code": op.type_code},
        )
        row = t.fetchone()
        if row:
            type_name = row[0]
    customer_name = None
    if op.customer_id:
        c = await session.execute(select(Customer.name_client, Customer.firm_name).where(Customer.id == op.customer_id))
        row = c.scalar_one_or_none()
        if row:
            customer_name = row[0] or row[1]
    STATUS_RU = {"pending": "Ð’ Ð¾Ð¶Ð¸Ð´Ð°Ð½Ð¸Ð¸", "completed": "Ð’Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¾", "cancelled": "ÐžÑ‚Ð¼ÐµÐ½ÐµÐ½Ð¾", "canceled": "ÐžÑ‚Ð¼ÐµÐ½ÐµÐ½Ð¾"}
    st = (op.status or "").strip().lower()
    return {
        "id": str(op.id),
        "operation_number": op.operation_number,
        "operation_date": op.operation_date.isoformat() if op.operation_date else None,
        "completed_date": op.completed_date.isoformat() if op.completed_date else None,
        "type_code": op.type_code,
        "type_name": type_name,
        "status": op.status,
        "status_name_ru": STATUS_RU.get(st, op.status or ""),
        "warehouse_from": op.warehouse_from,
        "warehouse_to": op.warehouse_to,
        "product_code": op.product_code,
        "quantity": op.quantity,
        "amount": float(op.amount) if op.amount else None,
        "payment_type_code": op.payment_type_code,
        "customer_id": op.customer_id,
        "customer_name": customer_name,
        "order_id": op.order_id,
        "comment": op.comment,
        "created_by": op.created_by,
        "expeditor_login": op.expeditor_login,
        "cashier_login": op.cashier_login,
        "storekeeper_login": op.storekeeper_login,
    }


class OperationUpdate(BaseModel):
    operation_date: datetime | None = None
    completed_date: datetime | None = None
    type_code: str | None = None
    warehouse_from: str | None = None
    warehouse_to: str | None = None
    product_code: str | None = None
    quantity: int | None = None
    amount: float | None = None
    payment_type_code: str | None = None
    customer_id: int | None = None
    order_id: int | None = None
    comment: str | None = None
    status: str | None = None
    expeditor_login: str | None = None
    cashier_login: str | None = None
    storekeeper_login: str | None = None


class OperationCreate(BaseModel):
    type_code: str
    operation_date: datetime | None = None
    warehouse_from: str | None = None
    warehouse_to: str | None = None
    product_code: str | None = None
    quantity: int | None = None
    amount: float | None = None
    payment_type_code: str | None = None
    customer_id: int | None = None
    order_id: int | None = None
    comment: str | None = None
    status: str | None = "pending"
    expeditor_login: str | None = None
    cashier_login: str | None = None
    storekeeper_login: str | None = None

    @field_validator("operation_date", mode="before")
    @classmethod
    def _default_operation_date(cls, value: datetime | None) -> datetime:
        return value or datetime.now(timezone.utc)

    @field_validator("quantity")
    @classmethod
    def _validate_quantity(cls, value: int | None) -> int | None:
        if value is not None and value < 0:
            raise ValueError("quantity must be >= 0")
        return value


@router.patch("/operations/{operation_id}")
async def update_operation(
    operation_id: UUID,
    body: OperationUpdate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Ð˜Ð·Ð¼ÐµÐ½Ð¸Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ. Ð¢Ð¾Ð»ÑŒÐºÐ¾ admin."""
    result = await session.execute(select(Operation).where(Operation.id == operation_id))
    op = result.scalar_one_or_none()
    if not op:
        raise HTTPException(status_code=404, detail="Operation not found")
    if body.operation_date is not None:
        op.operation_date = body.operation_date
    if body.completed_date is not None:
        op.completed_date = body.completed_date
    if body.type_code is not None:
        t = await session.execute(
            text('SELECT 1 FROM "Sales".operation_types WHERE code = :code'),
            {"code": body.type_code},
        )
        if t.fetchone() is None:
            raise HTTPException(status_code=400, detail=f"Ð¢Ð¸Ð¿ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸ Â«{body.type_code}Â» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½")
        op.type_code = body.type_code
    if body.warehouse_from is not None:
        op.warehouse_from = body.warehouse_from or None
    if body.warehouse_to is not None:
        op.warehouse_to = body.warehouse_to or None
    if body.product_code is not None:
        op.product_code = body.product_code or None
    if body.quantity is not None:
        op.quantity = body.quantity
    if body.amount is not None:
        op.amount = body.amount
    if body.payment_type_code is not None:
        op.payment_type_code = body.payment_type_code or None
    if body.customer_id is not None:
        op.customer_id = body.customer_id
    if body.order_id is not None:
        op.order_id = body.order_id
    if body.comment is not None:
        op.comment = body.comment
    if body.status is not None:
        op.status = body.status
    if body.expeditor_login is not None:
        op.expeditor_login = body.expeditor_login or None
    if body.cashier_login is not None:
        op.cashier_login = body.cashier_login or None
    if body.storekeeper_login is not None:
        op.storekeeper_login = body.storekeeper_login or None
    op.updated_at = datetime.now(timezone.utc)
    await session.commit()
    await session.refresh(op)
    return {"id": str(op.id), "operation_number": op.operation_number, "message": "updated"}


@router.post("/operations", response_model=EntityModel | list[EntityModel])
async def create_operation(
    body: OperationCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ. ÐÐ¾Ð¼ÐµÑ€ Ð³ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÑ‚ÑÑ Ð² Ð‘Ð” (generate_operation_number). Ð¢Ð¾Ð»ÑŒÐºÐ¾ admin."""
    try:
        type_check = await session.execute(
            text('SELECT 1 FROM "Sales".operation_types WHERE code = :code'),
            {"code": body.type_code},
        )
        if type_check.fetchone() is None:
            raise HTTPException(status_code=400, detail=f"Ð¢Ð¸Ð¿ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸ Â«{body.type_code}Â» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½")
        if body.order_id is not None:
            order_result = await session.execute(select(Order).where(Order.order_no == body.order_id))
            if order_result.scalar_one_or_none() is None:
                raise HTTPException(status_code=400, detail=f"Ð—Ð°ÐºÐ°Ð· Ñ Ð½Ð¾Ð¼ÐµÑ€Ð¾Ð¼ {body.order_id} Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½")
        if body.product_code:
            prod_result = await session.execute(select(Product).where(Product.code == body.product_code.strip()))
            if prod_result.scalar_one_or_none() is None:
                raise HTTPException(status_code=400, detail=f"Ð¢Ð¾Ð²Ð°Ñ€ Ñ ÐºÐ¾Ð´Ð¾Ð¼ Â«{body.product_code}Â» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½")
        num_result = await session.execute(text('SELECT "Sales".generate_operation_number()'))
        operation_number = num_result.scalar() or f"OP-{datetime.utcnow().strftime('%Y-%m-%d')}-000001"
        op = Operation(
            operation_number=operation_number,
            type_code=body.type_code,
            operation_date=body.operation_date,
            warehouse_from=(body.warehouse_from or "").strip() or None,
            warehouse_to=(body.warehouse_to or "").strip() or None,
            product_code=body.product_code.strip() if body.product_code else None,
            quantity=body.quantity,
            amount=body.amount,
            payment_type_code=body.payment_type_code or None,
            customer_id=body.customer_id,
            order_id=body.order_id,
            comment=body.comment,
            status=body.status or "pending",
            created_by=user.login,
            expeditor_login=body.expeditor_login or None,
            cashier_login=body.cashier_login or None,
            storekeeper_login=body.storekeeper_login or None,
        )
        session.add(op)
        await session.commit()
        await session.refresh(op)
    except HTTPException:
        await session.rollback()
        raise
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ²ÑÐ·Ð¸ Ñ Ð´Ð°Ð½Ð½Ñ‹Ð¼Ð¸: Ð¿Ñ€Ð¾Ð²ÐµÑ€ÑŒÑ‚Ðµ Ð·Ð°ÐºÐ°Ð·, Ñ‚Ð¾Ð²Ð°Ñ€, ÑÐºÐ»Ð°Ð´ Ð¸Ð»Ð¸ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ.")
    except Exception as e:
        await session.rollback()
        err_msg = str(e).strip() if e else "ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ"
        raise HTTPException(status_code=400, detail="ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ ÑÐ¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ: " + err_msg[:200])
    return {"id": str(op.id), "operation_number": op.operation_number, "type_code": body.type_code, "message": "created"}


logger = logging.getLogger(__name__)


@router.get("/operations/{operation_type}/form-config", response_model=EntityModel | list[EntityModel])
async def get_operation_form_config(
    operation_type: str,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÐºÐ¾Ð½Ñ„Ð¸Ð³ÑƒÑ€Ð°Ñ†Ð¸ÑŽ Ñ„Ð¾Ñ€Ð¼Ñ‹ Ð´Ð»Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸ Ð¾Ð¿Ñ€ÐµÐ´ÐµÐ»Ñ‘Ð½Ð½Ð¾Ð³Ð¾ Ñ‚Ð¸Ð¿Ð°.
    Ð’Ð¾Ð·Ð²Ñ€Ð°Ñ‰Ð°ÐµÑ‚ ÑÐ¿Ð¸ÑÐ¾Ðº Ð¿Ð¾Ð»ÐµÐ¹: required, optional, hidden, readonly.
    """
    logger.info(f"[FORM-CONFIG] Ð—Ð°Ð¿Ñ€Ð¾Ñ ÐºÐ¾Ð½Ñ„Ð¸Ð³Ð° Ð´Ð»Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸: {operation_type}")
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÐºÐ¾Ð½Ñ„Ð¸Ð³ Ð¸Ð· Ð‘Ð”
    config_result = await session.execute(
        select(OperationConfig).where(
            OperationConfig.operation_type_code == operation_type,
            OperationConfig.active == True
        )
    )
    config = config_result.scalar_one_or_none()
    
    if not config:
        logger.warning(f"[FORM-CONFIG] ÐšÐ¾Ð½Ñ„Ð¸Ð³ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½ Ð´Ð»Ñ: {operation_type}")
        raise HTTPException(
            status_code=404,
            detail=f"Configuration for operation '{operation_type}' not found"
        )
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð½Ð°Ð·Ð²Ð°Ð½Ð¸Ðµ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸ Ð¸Ð· operation_types
    type_result = await session.execute(
        select(OperationType).where(OperationType.code == operation_type)
    )
    op_type = type_result.scalar_one_or_none()
    operation_name = op_type.name if op_type else operation_type
    
    # ÐŸÐ°Ñ€ÑÐ¸Ñ‚ÑŒ JSON Ð¸Ð· Ð‘Ð”
    try:
        required_fields = json.loads(config.required_fields) if config.required_fields else []
        optional_fields = json.loads(config.optional_fields) if config.optional_fields else []
        hidden_fields = json.loads(config.hidden_fields) if config.hidden_fields else []
        readonly_fields = json.loads(config.readonly_fields) if config.readonly_fields else []
    except json.JSONDecodeError as e:
        logger.error(f"[FORM-CONFIG] ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ð°Ñ€ÑÐ¸Ð½Ð³Ð° JSON: {e}")
        raise HTTPException(status_code=500, detail=f"Invalid JSON in config: {str(e)}")
    
    logger.info(f"[FORM-CONFIG] ÐšÐ¾Ð½Ñ„Ð¸Ð³ Ð·Ð°Ð³Ñ€ÑƒÐ¶ÐµÐ½: required={len(required_fields)}, optional={len(optional_fields)}")
    
    return {
        "operation_type": config.operation_type_code,
        "operation_name": operation_name,
        "description": config.description,
        "default_status": config.default_status,
        "required_fields": required_fields,
        "optional_fields": optional_fields,
        "hidden_fields": hidden_fields,
        "readonly_fields": readonly_fields,
    }


class OperationCreateFromConfig(BaseModel):
    """ÐœÐ¾Ð´ÐµÐ»ÑŒ Ð´Ð»Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸ Ð¸Ð· ÐºÐ¾Ð½Ñ„Ð¸Ð³Ð° (Ð´Ð¸Ð½Ð°Ð¼Ð¸Ñ‡ÐµÑÐºÐ¸Ðµ Ð¿Ð¾Ð»Ñ)."""
    warehouse_from: str | None = None
    warehouse_to: str | None = None
    product_code: str | None = None
    batch_code: str | None = None
    expiry_date: str | None = None  # Ð´Ð´.Ð¼Ð¼.Ð³Ð³Ð³Ð³ Ð¸Ð»Ð¸ YYYY-MM-DD Ð´Ð»Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð¿Ð°Ñ€Ñ‚Ð¸Ð¸
    quantity: int | None = None
    amount: float | None = None
    payment_type_code: str | None = None
    customer_id: int | None = None
    order_id: int | None = None
    comment: str | None = None
    expeditor_login: str | None = None
    cashier_login: str | None = None
    storekeeper_login: str | None = None
    related_operation_id: str | None = None


@router.post("/operations/{operation_type}/create", response_model=EntityModel | list[EntityModel])
async def create_operation_from_config(
    operation_type: str,
    body: OperationCreateFromConfig,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ Ñ Ð²Ð°Ð»Ð¸Ð´Ð°Ñ†Ð¸ÐµÐ¹ Ð¿Ð¾ ÐºÐ¾Ð½Ñ„Ð¸Ð³Ñƒ Ð¸Ð· operation_config.
    ÐŸÑ€Ð¾Ð²ÐµÑ€ÑÐµÑ‚ required_fields, Ð¸Ð³Ð½Ð¾Ñ€Ð¸Ñ€ÑƒÐµÑ‚ hidden_fields, Ð·Ð°Ð¿Ð¾Ð»Ð½ÑÐµÑ‚ readonly_fields Ð°Ð²Ñ‚Ð¾Ð¼Ð°Ñ‚Ð¸Ñ‡ÐµÑÐºÐ¸.
    """
    logger.info(f"=== CREATING OPERATION ===")
    logger.info(f"Operation type: {operation_type}")
    logger.info(f"Current user: {user.login}")
    logger.info(f"Incoming data: {body.model_dump()}")
    
    # Ð¨ÐÐ“ 1: ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÐºÐ¾Ð½Ñ„Ð¸Ð³
    logger.info(f"[STEP 1] ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ ÐºÐ¾Ð½Ñ„Ð¸Ð³ Ð´Ð»Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸: {operation_type}")
    
    config_result = await session.execute(
        select(OperationConfig).where(
            OperationConfig.operation_type_code == operation_type,
            OperationConfig.active == True
        )
    )
    config = config_result.scalar_one_or_none()
    
    if not config:
        logger.error(f"[STEP 1] ÐšÐ¾Ð½Ñ„Ð¸Ð³ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½ Ð´Ð»Ñ: {operation_type}")
        raise HTTPException(
            status_code=404,
            detail=f"Configuration for '{operation_type}' not found"
        )
    await _ensure_user_can_create_operation_type(session, user, operation_type)
    logger.info(f"âœ“ ÐšÐ¾Ð½Ñ„Ð¸Ð³ Ð½Ð°Ð¹Ð´ÐµÐ½")
    
    # Ð¨ÐÐ“ 2: ÐŸÐ°Ñ€ÑÐ¸Ñ‚ÑŒ required_fields Ð¸Ð· ÐºÐ¾Ð½Ñ„Ð¸Ð³Ð°
    logger.info(f"[STEP 2] ÐŸÑ€Ð¾Ð²ÐµÑ€ÑÐµÐ¼ Ð¾Ð±ÑÐ·Ð°Ñ‚ÐµÐ»ÑŒÐ½Ñ‹Ðµ Ð¿Ð¾Ð»Ñ")
    
    try:
        required_fields = json.loads(config.required_fields) if config.required_fields else []
        optional_fields = json.loads(config.optional_fields) if config.optional_fields else []
        hidden_fields = json.loads(config.hidden_fields) if config.hidden_fields else []
    except json.JSONDecodeError as e:
        logger.error(f"[STEP 2] ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ð°Ñ€ÑÐ¸Ð½Ð³Ð° JSON: {e}")
        raise HTTPException(status_code=500, detail=f"Invalid JSON in config: {str(e)}")
    
    logger.info(f"  Required: {required_fields}")
    logger.info(f"  Optional: {optional_fields}")
    logger.info(f"  Hidden: {hidden_fields}")
    
    # Ð¨ÐÐ“ 3: ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ, Ñ‡Ñ‚Ð¾ Ð²ÑÐµ required_fields Ð·Ð°Ð¿Ð¾Ð»Ð½ÐµÐ½Ñ‹
    logger.info(f"[STEP 3] Ð’Ð°Ð»Ð¸Ð´Ð¸Ñ€ÑƒÐµÐ¼ Ð·Ð°Ð¿Ð¾Ð»Ð½ÐµÐ½Ð½Ð¾ÑÑ‚ÑŒ required_fields")
    
    data = body.model_dump(exclude_none=True)
    errors = {}
    
    for field in required_fields:
        # Ð¡Ð¿ÐµÑ†Ð¸Ð°Ð»ÑŒÐ½Ñ‹Ðµ Ð¿Ð¾Ð»Ñ, ÐºÐ¾Ñ‚Ð¾Ñ€Ñ‹Ðµ Ð·Ð°Ð¿Ð¾Ð»Ð½ÑÑŽÑ‚ÑÑ Ð°Ð²Ñ‚Ð¾Ð¼Ð°Ñ‚Ð¸Ñ‡ÐµÑÐºÐ¸
        if field in ['created_by', 'operation_date', 'operation_number', 'status', 'type_code']:
            continue
        
        # ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ, Ñ‡Ñ‚Ð¾ Ð¿Ð¾Ð»Ðµ Ð·Ð°Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¾
        if field not in data or data[field] is None or (isinstance(data[field], str) and data[field].strip() == ''):
            errors[field] = "This field is required"
            logger.info(f"  âŒ {field}: MISSING")
        else:
            logger.info(f"  âœ“ {field}: OK")
    
    # Ð•ÑÐ»Ð¸ ÐµÑÑ‚ÑŒ Ð¾ÑˆÐ¸Ð±ÐºÐ¸ Ð²Ð°Ð»Ð¸Ð´Ð°Ñ†Ð¸Ð¸ â€” Ð²ÐµÑ€Ð½ÑƒÑ‚ÑŒ Ð¸Ñ…
    if errors:
        logger.warning(f"\nâŒ ÐžÐ¨Ð˜Ð‘ÐšÐ˜ Ð’ÐÐ›Ð˜Ð”ÐÐ¦Ð˜Ð˜: {errors}")
        raise HTTPException(status_code=400, detail={"errors": errors})
    
    # Ð¨ÐÐ“ 4: ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ, Ñ‡Ñ‚Ð¾ Ð½Ð¸ÐºÐ°ÐºÐ¸Ðµ hidden_fields Ð½Ðµ Ð¿ÐµÑ€ÐµÐ´Ð°Ð½Ñ‹
    logger.info(f"\n[STEP 4] ÐŸÑ€Ð¾Ð²ÐµÑ€ÑÐµÐ¼, Ñ‡Ñ‚Ð¾ ÑÐºÑ€Ñ‹Ñ‚Ñ‹Ðµ Ð¿Ð¾Ð»Ñ Ð½Ðµ Ð¿ÐµÑ€ÐµÐ´Ð°Ð½Ñ‹")
    
    for field in hidden_fields:
        if field in data:
            logger.warning(f"  âš ï¸ Ð’ÐÐ˜ÐœÐÐÐ˜Ð•: ÐŸÐ¾Ð»Ðµ {field} Ð´Ð¾Ð»Ð¶Ð½Ð¾ Ð±Ñ‹Ñ‚ÑŒ ÑÐºÑ€Ñ‹Ñ‚Ð¾, Ð½Ð¾ Ð¿ÐµÑ€ÐµÐ´Ð°Ð½Ð¾ Ð² data!")
            del data[field]
    
    logger.info(f"âœ“ Ð¡ÐºÑ€Ñ‹Ñ‚Ñ‹Ðµ Ð¿Ð¾Ð»Ñ Ð¾Ñ‡Ð¸Ñ‰ÐµÐ½Ñ‹")
    
    # Ð¨ÐÐ“ 5: Ð—Ð°Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÑŒ readonly_fields Ð°Ð²Ñ‚Ð¾Ð¼Ð°Ñ‚Ð¸Ñ‡ÐµÑÐºÐ¸
    logger.info(f"\n[STEP 5] Ð—Ð°Ð¿Ð¾Ð»Ð½ÑÐµÐ¼ readonly_fields")
    
    # Ð“ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÐ¼ Ð½Ð¾Ð¼ÐµÑ€ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸
    num_result = await session.execute(text('SELECT "Sales".generate_operation_number()'))
    operation_number = num_result.scalar() or f"OP-{datetime.now(timezone.utc).strftime('%Y-%m-%d')}-000001"
    logger.info(f"  Generated operation_number: {operation_number}")
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ ÑÑ‚Ð°Ñ‚ÑƒÑ Ð¸Ð· ÐºÐ¾Ð½Ñ„Ð¸Ð³Ð° (Ð´Ð»Ñ Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ â€” Ð²ÑÐµÐ³Ð´Ð° completed)
    default_status = config.default_status
    if operation_type == "delivery":
        default_status = "completed"
    logger.info(f"  default_status: {default_status}")
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ñ‚ÐµÐºÑƒÑ‰ÐµÐ³Ð¾ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ
    created_by = user.login
    logger.info(f"  created_by: {created_by}")
    
    # Ð”Ð»Ñ Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸: Ð¿Ñ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ, Ñ‡Ñ‚Ð¾ Ð·Ð°ÐºÐ°Ð· Ð½Ðµ ÑƒÐ¶Ðµ Ð´Ð¾ÑÑ‚Ð°Ð²Ð»ÐµÐ½, Ð¸ Ñ‡Ñ‚Ð¾ Ð¾ÑÑ‚Ð°Ñ‚ÐºÐ¾Ð² Ñ…Ð²Ð°Ñ‚Ð°ÐµÑ‚
    if operation_type == "delivery":
        order_id_val = data.get("order_id")
        if order_id_val is not None:
            order_row = await session.execute(select(Order).where(Order.order_no == order_id_val))
            order_obj = order_row.scalar_one_or_none()
            if order_obj and (order_obj.status_code or "").strip().lower() == "completed":
                raise HTTPException(
                    status_code=400,
                    detail="Ð—Ð°ÐºÐ°Ð· ÑƒÐ¶Ðµ Ð´Ð¾ÑÑ‚Ð°Ð²Ð»ÐµÐ½. ÐÐµÐ»ÑŒÐ·Ñ ÑÐ¾Ð·Ð´Ð°Ñ‚ÑŒ ÐµÑ‰Ñ‘ Ð¾Ð´Ð½Ñƒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ Ð´Ð¾ÑÑ‚Ð°Ð²ÐºÐ¸ Ð¿Ð¾ ÑÑ‚Ð¾Ð¼Ñƒ Ð·Ð°ÐºÐ°Ð·Ñƒ.",
                )
        wh_from = data.get("warehouse_from")
        pc = data.get("product_code")
        bc = data.get("batch_code")
        qty = data.get("quantity") or 0
        if wh_from and pc and bc is not None and qty > 0:
            try:
                stock_r = await session.execute(
                    text('''
                        SELECT COALESCE(total_qty, 0)::int FROM "Sales".v_warehouse_stock
                        WHERE warehouse_code = :wh AND product_code = :pc AND batch_code = :bc
                    '''),
                    {"wh": wh_from, "pc": pc, "bc": bc},
                )
                stock_row = stock_r.fetchone()
                available = int(stock_row[0]) if stock_row and stock_row[0] is not None else 0
                if available < qty:
                    raise HTTPException(
                        status_code=400,
                        detail=f"ÐÐµÐ´Ð¾ÑÑ‚Ð°Ñ‚Ð¾Ñ‡Ð½Ð¾ Ð¾ÑÑ‚Ð°Ñ‚ÐºÐ¾Ð² Ð½Ð° ÑÐºÐ»Ð°Ð´Ðµ: Ð´Ð¾ÑÑ‚ÑƒÐ¿Ð½Ð¾ {available}, Ð·Ð°Ð¿Ñ€Ð¾ÑˆÐµÐ½Ð¾ {qty}.",
                    )
            except HTTPException:
                raise
            except Exception as stock_check_error:
                logger.warning(
                    "Delivery stock pre-check skipped due to error: %s",
                    stock_check_error,
                )
    
    # Ð¨ÐÐ“ 6: Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ Ð² Ð‘Ð”
    logger.info(f"\n[STEP 6] Ð¡Ð¾Ð·Ð´Ð°Ñ‘Ð¼ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ Ð² Ð‘Ð”")
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð¸Ð»Ð¸ ÑÐ¾Ð·Ð´Ð°Ñ‚ÑŒ batch Ð¿Ð¾ batch_code Ð¸ product_code
    batch_id = None
    if data.get("batch_code") and data.get("product_code"):
        batch_result = await session.execute(
            select(Batch).where(
                Batch.batch_code == data["batch_code"],
                Batch.product_code == data["product_code"]
            )
        )
        batch = batch_result.scalar_one_or_none()
        if batch:
            batch_id = batch.id
        else:
            expiry_date_parsed = None
            if data.get("expiry_date"):
                raw = data["expiry_date"].strip()
                for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                    try:
                        expiry_date_parsed = datetime.strptime(raw, fmt).date()
                        break
                    except ValueError:
                        continue
            new_batch = Batch(
                product_code=data["product_code"],
                batch_code=data["batch_code"],
                expiry_date=expiry_date_parsed,
            )
            session.add(new_batch)
            await session.flush()
            batch_id = new_batch.id
            logger.info(f"  Ð¡Ð¾Ð·Ð´Ð°Ð½Ð° Ð½Ð¾Ð²Ð°Ñ Ð¿Ð°Ñ€Ñ‚Ð¸Ñ: {data['batch_code']}, expiry_date={expiry_date_parsed}")

    await _ensure_sufficient_batch_stock(
        session=session,
        operation_type=operation_type,
        warehouse_from=data.get("warehouse_from"),
        product_code=data.get("product_code"),
        batch_code=data.get("batch_code"),
        qty=data.get("quantity"),
    )

    if operation_type == "promotional_sample":
        data["amount"] = 0
    
    op = Operation(
        operation_number=operation_number,
        type_code=operation_type,
        status=default_status,
        operation_date=datetime.now(timezone.utc),
        created_by=created_by,
        warehouse_from=data.get("warehouse_from"),
        warehouse_to=data.get("warehouse_to"),
        product_code=data.get("product_code"),
        batch_id=batch_id,
        quantity=data.get("quantity"),
        amount=data.get("amount"),
        payment_type_code=data.get("payment_type_code"),
        customer_id=data.get("customer_id"),
        order_id=data.get("order_id"),
        comment=data.get("comment"),
        expeditor_login=data.get("expeditor_login"),
        cashier_login=data.get("cashier_login"),
        storekeeper_login=data.get("storekeeper_login"),
        related_operation_id=UUID(data["related_operation_id"]) if data.get("related_operation_id") else None,
    )
    
    session.add(op)
    await session.flush()
    if operation_type == "payment_receipt_from_customer":
        num2 = (await session.execute(text('SELECT "Sales".generate_operation_number()'))).scalar() or "OP-000000"
        handover = Operation(
            operation_number=num2,
            type_code="cash_handover_from_expeditor",
            status="pending",
            operation_date=datetime.now(timezone.utc),
            created_by=op.created_by,
            amount=op.amount,
            expeditor_login=op.expeditor_login,
            customer_id=op.customer_id,
            order_id=op.order_id,
            related_operation_id=op.id,
        )
        session.add(handover)
        logger.info(f"  ÐÐ²Ñ‚Ð¾ÑÐ¾Ð·Ð´Ð°Ð½Ð° cash_handover_from_expeditor: {num2}")
    try:
        await session.commit()
        await session.refresh(op)
    except IntegrityError as e:
        await session.rollback()
        logger.error(f"[STEP 6] ÐžÑˆÐ¸Ð±ÐºÐ° IntegrityError: {e}")
        raise HTTPException(status_code=400, detail="ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ²ÑÐ·Ð¸ Ñ Ð´Ð°Ð½Ð½Ñ‹Ð¼Ð¸: Ð¿Ñ€Ð¾Ð²ÐµÑ€ÑŒÑ‚Ðµ Ð·Ð°ÐºÐ°Ð·, Ñ‚Ð¾Ð²Ð°Ñ€, ÑÐºÐ»Ð°Ð´ Ð¸Ð»Ð¸ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ.")
    except Exception as e:
        await session.rollback()
        logger.error(f"[STEP 6] ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸: {e}")
        err_msg = str(e).strip() if e else "ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ"
        raise HTTPException(status_code=400, detail="ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ ÑÐ¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ: " + err_msg[:200])
    
    logger.info(f"âœ“ ÐžÐ¿ÐµÑ€Ð°Ñ†Ð¸Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð°: {op.operation_number}")
    logger.info(f"=== END ===")
    
    # Ð¨ÐÐ“ 7: Ð’ÐµÑ€Ð½ÑƒÑ‚ÑŒ Ñ€ÐµÐ·ÑƒÐ»ÑŒÑ‚Ð°Ñ‚
    return {
        "status": "success",
        "operation_id": str(op.id),
        "operation_number": op.operation_number,
        "type": operation_type,
        "default_status": op.status,
        "message": f"Operation '{operation_type}' created successfully"
    }


class WarehouseReceiptBatchItem(BaseModel):
    """ÐžÐ´Ð½Ð° Ð¿Ð¾Ð·Ð¸Ñ†Ð¸Ñ Ñ‚Ð¾Ð²Ð°Ñ€Ð° Ð´Ð»Ñ Ð¿Ñ€Ð¸Ñ…Ð¾Ð´Ð° Ð½Ð° ÑÐºÐ»Ð°Ð´."""
    product_code: str
    expiry_date: str  # Ð´Ð´.Ð¼Ð¼.Ð³Ð³Ð³Ð³ Ð¸Ð»Ð¸ YYYY-MM-DD
    quantity: int
    batch_code: str  # Ð°Ð²Ñ‚Ð¾Ð³ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÑ‚ÑÑ, Ð½Ð¾ Ð¿ÐµÑ€ÐµÐ´Ð°Ñ‘Ñ‚ÑÑ Ð´Ð»Ñ Ð¿Ñ€Ð¾Ð²ÐµÑ€ÐºÐ¸


class WarehouseReceiptBatchCreate(BaseModel):
    """Ð¡Ð¾Ð·Ð´Ð°Ð½Ð¸Ðµ Ð½ÐµÑÐºÐ¾Ð»ÑŒÐºÐ¸Ñ… Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹ Ð¿Ñ€Ð¸Ñ…Ð¾Ð´Ð° Ð½Ð° ÑÐºÐ»Ð°Ð´."""
    warehouse_to: str
    items: list[WarehouseReceiptBatchItem]
    comment: str | None = None


@router.post("/operations/warehouse_receipt/create-batch", response_model=EntityModel | list[EntityModel])
async def create_warehouse_receipt_batch(
    body: WarehouseReceiptBatchCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð½ÐµÑÐºÐ¾Ð»ÑŒÐºÐ¾ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹ Ð¿Ñ€Ð¸Ñ…Ð¾Ð´Ð° Ð½Ð° ÑÐºÐ»Ð°Ð´ Ð·Ð° Ñ€Ð°Ð· (Ð´Ð»Ñ ÐºÐ°Ð¶Ð´Ð¾Ð³Ð¾ Ñ‚Ð¾Ð²Ð°Ñ€Ð° Ð¾Ñ‚Ð´ÐµÐ»ÑŒÐ½Ð°Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ñ).
    batch_code Ð³ÐµÐ½ÐµÑ€Ð¸Ñ€ÑƒÐµÑ‚ÑÑ Ð°Ð²Ñ‚Ð¾Ð¼Ð°Ñ‚Ð¸Ñ‡ÐµÑÐºÐ¸ ÐºÐ°Ðº {product_code}_{DDMMYYYY} Ð½Ð° Ð¾ÑÐ½Ð¾Ð²Ðµ expiry_date.
    """
    logger.info(f"=== CREATING WAREHOUSE RECEIPT BATCH ===")
    logger.info(f"Warehouse: {body.warehouse_to}, Items: {len(body.items)}")
    logger.info(f"Current user: {user.login}")
    await _ensure_user_can_create_operation_type(session, user, "warehouse_receipt")
    
    if not body.items:
        raise HTTPException(status_code=400, detail="Ð¡Ð¿Ð¸ÑÐ¾Ðº Ñ‚Ð¾Ð²Ð°Ñ€Ð¾Ð² Ð¿ÑƒÑÑ‚")
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÐºÐ¾Ð½Ñ„Ð¸Ð³ Ð´Ð»Ñ warehouse_receipt
    config_result = await session.execute(
        select(OperationConfig).where(
            OperationConfig.operation_type_code == 'warehouse_receipt',
            OperationConfig.active == True
        )
    )
    config = config_result.scalar_one_or_none()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration for 'warehouse_receipt' not found")
    
    default_status = config.default_status
    created_operations = []
    errors = []
    
    for idx, item in enumerate(body.items):
        try:
            # ÐŸÐ°Ñ€ÑÐ¸Ñ‚ÑŒ expiry_date
            expiry_date_parsed = None
            raw = item.expiry_date.strip()
            for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                try:
                    expiry_date_parsed = datetime.strptime(raw, fmt).date()
                    break
                except ValueError:
                    continue
            
            if not expiry_date_parsed:
                errors.append(f"ÐŸÐ¾Ð·Ð¸Ñ†Ð¸Ñ {idx + 1}: Ð½ÐµÐ²ÐµÑ€Ð½Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚ Ð´Ð°Ñ‚Ñ‹ '{item.expiry_date}'")
                continue
            
            # ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ batch_code (Ð´Ð¾Ð»Ð¶ÐµÐ½ Ð±Ñ‹Ñ‚ÑŒ product_code_DDMMYYYY)
            expected_batch_code = f"{item.product_code}_{expiry_date_parsed.strftime('%d%m%Y')}"
            if item.batch_code != expected_batch_code:
                logger.warning(f"  ÐŸÐ¾Ð·Ð¸Ñ†Ð¸Ñ {idx + 1}: batch_code Ð½Ðµ ÑÐ¾Ð²Ð¿Ð°Ð´Ð°ÐµÑ‚. ÐžÐ¶Ð¸Ð´Ð°Ð»Ð¾ÑÑŒ: {expected_batch_code}, Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¾: {item.batch_code}")
                # Ð˜ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐµÐ¼ Ð¾Ð¶Ð¸Ð´Ð°ÐµÐ¼Ñ‹Ð¹ ÐºÐ¾Ð´
                batch_code = expected_batch_code
            else:
                batch_code = item.batch_code
            
            # ÐÐ°Ð¹Ñ‚Ð¸ Ð¸Ð»Ð¸ ÑÐ¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¿Ð°Ñ€Ñ‚Ð¸ÑŽ
            batch_result = await session.execute(
                select(Batch).where(
                    Batch.batch_code == batch_code,
                    Batch.product_code == item.product_code
                )
            )
            batch = batch_result.scalar_one_or_none()
            
            if not batch:
                new_batch = Batch(
                    product_code=item.product_code,
                    batch_code=batch_code,
                    expiry_date=expiry_date_parsed,
                )
                session.add(new_batch)
                await session.flush()
                batch_id = new_batch.id
                logger.info(f"  Ð¡Ð¾Ð·Ð´Ð°Ð½Ð° Ð¿Ð°Ñ€Ñ‚Ð¸Ñ: {batch_code}")
            else:
                batch_id = batch.id
            
            # Ð“ÐµÐ½ÐµÑ€Ð¸Ñ€Ð¾Ð²Ð°Ñ‚ÑŒ Ð½Ð¾Ð¼ÐµÑ€ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸
            num_result = await session.execute(text('SELECT "Sales".generate_operation_number()'))
            operation_number = num_result.scalar() or f"OP-{datetime.now(timezone.utc).strftime('%Y-%m-%d')}-000001"
            
            # Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ
            op = Operation(
                operation_number=operation_number,
                type_code='warehouse_receipt',
                status=default_status,
                operation_date=datetime.now(timezone.utc),
                created_by=user.login,
                warehouse_to=body.warehouse_to,
                product_code=item.product_code,
                batch_id=batch_id,
                quantity=item.quantity,
                comment=body.comment,
            )
            session.add(op)
            await session.flush()
            created_operations.append({
                "operation_number": op.operation_number,
                "product_code": item.product_code,
                "quantity": item.quantity,
            })
            logger.info(f"  Ð¡Ð¾Ð·Ð´Ð°Ð½Ð° Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ñ: {op.operation_number} Ð´Ð»Ñ Ñ‚Ð¾Ð²Ð°Ñ€Ð° {item.product_code}")
            
        except Exception as e:
            logger.error(f"  ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ð¸ Ð¿Ð¾Ð·Ð¸Ñ†Ð¸Ð¸ {idx + 1}: {e}")
            errors.append(f"ÐŸÐ¾Ð·Ð¸Ñ†Ð¸Ñ {idx + 1}: {str(e)[:100]}")
            continue
    
    if errors and not created_operations:
        await session.rollback()
        raise HTTPException(status_code=400, detail={"errors": errors})
    
    try:
        await session.commit()
        logger.info(f"âœ“ Ð¡Ð¾Ð·Ð´Ð°Ð½Ð¾ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹: {len(created_operations)}")
        logger.info(f"=== END ===")
    except Exception as e:
        await session.rollback()
        logger.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° ÐºÐ¾Ð¼Ð¼Ð¸Ñ‚Ð°: {e}")
        raise HTTPException(status_code=400, detail=f"ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ ÑÐ¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸: {str(e)[:200]}")
    
    return {
        "status": "success",
        "operations_count": len(created_operations),
        "operations": created_operations,
        "errors": errors if errors else None,
        "message": f"Ð¡Ð¾Ð·Ð´Ð°Ð½Ð¾ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¹: {len(created_operations)}"
    }


class AllocationCreate(BaseModel):
    """ÐœÐ¾Ð´ÐµÐ»ÑŒ Ð´Ð»Ñ Ð²Ñ‹Ð´Ð°Ñ‡Ð¸ Ñ‚Ð¾Ð²Ð°Ñ€Ð° ÑÐºÑÐ¿ÐµÐ´Ð¸Ñ‚Ð¾Ñ€Ñƒ (allocation)."""
    warehouse_from: str
    warehouse_to: str
    product_code: str
    batch_code: str
    quantity: int
    expeditor_login: str
    comment: str | None = None


@router.post("/operations/allocation", response_model=EntityModel | list[EntityModel])
async def create_allocation_operation(
    body: AllocationCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¾Ð´Ð½Ñƒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ Â«allocationÂ» (Ð²Ñ‹Ð´Ð°Ñ‡Ð° ÑÐºÑÐ¿ÐµÐ´Ð¸Ñ‚Ð¾Ñ€Ñƒ).

    Ð˜ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐµÑ‚ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÑŽÑ‰ÑƒÑŽ Ð¿Ð°Ñ€Ñ‚Ð¸ÑŽ (batch) Ð¿Ð¾ product_code + batch_code,
    Ð½Ðµ ÑÐ¾Ð·Ð´Ð°Ñ‘Ñ‚ Ð½Ð¾Ð²Ñ‹Ðµ Ð¿Ð°Ñ€Ñ‚Ð¸Ð¸.
    """
    await _ensure_user_can_create_operation_type(session, user, "allocation")

    # ÐÐ°Ð¹Ñ‚Ð¸ Ð¿Ð°Ñ€Ñ‚Ð¸ÑŽ Ð¿Ð¾ product_code + batch_code
    batch_result = await session.execute(
        select(Batch).where(
            Batch.product_code == body.product_code.strip(),
            Batch.batch_code == body.batch_code.strip(),
        )
    )
    batch = batch_result.scalar_one_or_none()
    if not batch:
        raise HTTPException(
            status_code=400,
            detail=f"ÐŸÐ°Ñ€Ñ‚Ð¸Ñ Ñ ÐºÐ¾Ð´Ð¾Ð¼ Â«{body.batch_code}Â» Ð´Ð»Ñ Ñ‚Ð¾Ð²Ð°Ñ€Ð° Â«{body.product_code}Â» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½Ð°",
        )

    # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÐºÐ¾Ð½Ñ„Ð¸Ð³ Ð´Ð»Ñ allocation, Ñ‡Ñ‚Ð¾Ð±Ñ‹ Ð²Ð·ÑÑ‚ÑŒ default_status
    config_result = await session.execute(
        select(OperationConfig).where(
            OperationConfig.operation_type_code == "allocation",
            OperationConfig.active == True,
        )
    )
    config = config_result.scalar_one_or_none()
    default_status = config.default_status if config else "completed"

    # Ð¡Ð³ÐµÐ½ÐµÑ€Ð¸Ñ€Ð¾Ð²Ð°Ñ‚ÑŒ Ð½Ð¾Ð¼ÐµÑ€ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸
    num_result = await session.execute(text('SELECT "Sales".generate_operation_number()'))
    operation_number = num_result.scalar() or f"OP-{datetime.utcnow().strftime('%Y-%m-%d')}-000001"

    op = Operation(
        operation_number=operation_number,
        type_code="allocation",
        status=default_status,
        operation_date=datetime.now(timezone.utc),
        created_by=user.login,
        warehouse_from=body.warehouse_from.strip(),
        warehouse_to=body.warehouse_to.strip(),
        product_code=body.product_code.strip(),
        batch_id=batch.id,
        quantity=body.quantity,
        comment=body.comment,
        expeditor_login=body.expeditor_login.strip(),
    )

    session.add(op)
    try:
        await session.commit()
        await session.refresh(op)
    except IntegrityError as e:
        await session.rollback()
        logger.error(f"[ALLOCATION] IntegrityError: {e}")
        raise HTTPException(
            status_code=400,
            detail="ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ²ÑÐ·Ð¸ Ñ Ð´Ð°Ð½Ð½Ñ‹Ð¼Ð¸: Ð¿Ñ€Ð¾Ð²ÐµÑ€ÑŒÑ‚Ðµ ÑÐºÐ»Ð°Ð´Ñ‹, Ñ‚Ð¾Ð²Ð°Ñ€, Ð¿Ð°Ñ€Ñ‚Ð¸ÑŽ Ð¸Ð»Ð¸ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ.",
        )
    except Exception as e:
        await session.rollback()
        logger.error(f"[ALLOCATION] ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸Ð¸: {e}")
        err_msg = str(e).strip() if e else "ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ"
        raise HTTPException(status_code=400, detail="ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ ÑÐ¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ Ð¾Ð¿ÐµÑ€Ð°Ñ†Ð¸ÑŽ: " + err_msg[:200])

    return {
        "status": "success",
        "operation_id": str(op.id),
        "operation_number": op.operation_number,
        "type": op.type_code,
        "message": "ÐžÐ¿ÐµÑ€Ð°Ñ†Ð¸Ñ Ð²Ñ‹Ð´Ð°Ñ‡Ð¸ ÑÐºÑÐ¿ÐµÐ´Ð¸Ñ‚Ð¾Ñ€Ñƒ ÑÐ¾Ð·Ð´Ð°Ð½Ð°",
    }




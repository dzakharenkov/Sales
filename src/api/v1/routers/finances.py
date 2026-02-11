"""
GET /finances/ledger — финансовый реестр из VIEW v_financial_ledger (ТЗ).
GET /finances/pending-handovers — ожидающие передачи от экспедиторов (cash_handover_from_expeditor, status=pending).
POST /finances/accept-handover — принять передачу: cash_handover → completed, автосоздание cash_receipt.
GET /finances/cash-received — принятые деньги за период (cash_receipt).
GET /finances/orders-for-confirmation — заказы для подтверждения оплаты кассиром.
"""
import io
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook

from src.database.connection import get_db_session
from src.core.deps import get_current_user
from src.database.models import User, Operation

router = APIRouter()


def _parse_date_to_iso(s: str | None) -> str | None:
    """Приводит дату к YYYY-MM-DD. Поддерживает Y-m-d и d.m.Y."""
    if not s or not (s := s.strip()):
        return None
    s = s[:10]
    if len(s) == 10:
        if "-" in s and s[4] == "-" and s[7] == "-":
            return s
        if "." in s:
            parts = s.split(".")
            if len(parts) == 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if 1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100:
                        return f"{y:04d}-{m:02d}-{d:02d}"
                except (ValueError, TypeError):
                    pass
    return None


class AcceptHandoverBody(BaseModel):
    handover_operation_id: str
    cashier_login: str
    amount: float | None = None
    comment: str | None = None


class CashReceiptManualBody(BaseModel):
    """Ручное создание cash_receipt для безналичных переводов (без related_operation_id)."""
    amount: float
    payment_type_code: str  # bank_sum, card_sum и т.д.
    cashier_login: str
    customer_id: int | None = None
    order_id: int | None = None
    comment: str | None = None


@router.get("/pending-handovers")
async def get_pending_handovers(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Ожидающие передачи наличных от экспедиторов (cash_handover_from_expeditor, status=pending)."""
    try:
        q = """
        SELECT o.id, o.operation_number, o.warehouse_from, o.warehouse_to, o.amount,
               o.expeditor_login, o.operation_date, o.comment, o.order_id, o.customer_id,
               u.fio AS expeditor_fio,
               (c.name_client || COALESCE(' ' || c.firm_name, '')) AS customer_name
        FROM "Sales".operations o
        LEFT JOIN "Sales".users u ON o.expeditor_login = u.login
        LEFT JOIN "Sales".customers c ON o.customer_id = c.id
        WHERE o.type_code = 'cash_handover_from_expeditor' AND o.status = 'pending'
        ORDER BY o.operation_date ASC
        """
        r = await session.execute(text(q))
        rows = r.fetchall()
        cols = [c for c in r.keys()]
        data = [dict(zip(cols, row)) for row in rows]
        for d in data:
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)[:200], "data": []}


@router.post("/accept-handover")
async def accept_handover(
    body: AcceptHandoverBody,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    Принять передачу наличных от экспедитора.
    Переводит cash_handover_from_expeditor в completed и автоматически создаёт cash_receipt.
    """
    try:
        handover_id = UUID(body.handover_operation_id)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Некорректный ID операции передачи")
    r = await session.execute(
        select(Operation).where(
            Operation.id == handover_id,
            Operation.type_code == "cash_handover_from_expeditor",
            Operation.status == "pending",
        )
    )
    handover = r.scalar_one_or_none()
    if not handover:
        raise HTTPException(
            status_code=400,
            detail="Операция передачи не найдена или уже принята",
        )
    amount = float(body.amount) if body.amount is not None else float(handover.amount or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Укажите положительную сумму")
    handover.status = "completed"
    num_result = await session.execute(text('SELECT "Sales".generate_operation_number()'))
    op_number = num_result.scalar() or "OP-000000"
    cash_receipt = Operation(
        operation_number=op_number,
        type_code="cash_receipt",
        status="completed",
        operation_date=datetime.now(timezone.utc),
        created_by=user.login,
        amount=amount,
        payment_type_code="cash_sum",
        cashier_login=body.cashier_login,
        expeditor_login=handover.expeditor_login,
        related_operation_id=handover.id,
        comment=body.comment,
    )
    session.add(cash_receipt)
    await session.commit()
    await session.refresh(cash_receipt)
    return {
        "success": True,
        "handover_id": str(handover.id),
        "cash_receipt_id": str(cash_receipt.id),
        "cash_receipt_number": cash_receipt.operation_number,
        "amount": amount,
        "message": "Наличные приняты в кассу",
    }


@router.post("/cash-receipt-manual")
async def create_cash_receipt_manual(
    body: CashReceiptManualBody,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    Ручное создание cash_receipt для безналичных переводов (банк, карта и т.д.).
    Без related_operation_id — источник: от клиента.
    """
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Сумма должна быть положительной")
    pt_check = await session.execute(
        text('SELECT 1 FROM "Sales".payment_type WHERE code = :c'),
        {"c": body.payment_type_code},
    )
    if pt_check.scalar() is None:
        raise HTTPException(status_code=400, detail=f"Тип оплаты «{body.payment_type_code}» не найден")
    num_result = await session.execute(text('SELECT "Sales".generate_operation_number()'))
    op_number = num_result.scalar() or "OP-000000"
    op = Operation(
        operation_number=op_number,
        type_code="cash_receipt",
        status="completed",
        operation_date=datetime.now(timezone.utc),
        created_by=user.login,
        amount=body.amount,
        payment_type_code=body.payment_type_code,
        cashier_login=body.cashier_login,
        customer_id=body.customer_id,
        order_id=body.order_id,
        comment=body.comment,
    )
    session.add(op)
    await session.commit()
    await session.refresh(op)
    return {
        "success": True,
        "id": str(op.id),
        "operation_number": op.operation_number,
        "amount": float(op.amount),
        "message": "Платёж зафиксирован в кассе",
    }


@router.get("/cash-received")
async def get_cash_received(
    date_from: str | None = Query(None, description="Дата с (YYYY-MM-DD)"),
    date_to: str | None = Query(None, description="Дата по (YYYY-MM-DD)"),
    tz: str = Query("Asia/Tashkent", description="Часовой пояс для интерпретации дат"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Принятые деньги в кассу за период (cash_receipt, completed)."""
    try:
        tz_safe = tz.replace("'", "").replace(";", "")[:50] if tz else "UTC"
        params = {"tz": tz_safe}
        date_expr = "(o.operation_date AT TIME ZONE :tz)::date"
        conditions = ["o.type_code = 'cash_receipt'", "o.status = 'completed'"]
        df_iso = _parse_date_to_iso(date_from)
        dt_iso = _parse_date_to_iso(date_to)
        if df_iso:
            conditions.append(f"{date_expr} >= :date_from")
            params["date_from"] = df_iso
        if dt_iso:
            conditions.append(f"{date_expr} <= :date_to")
            params["date_to"] = dt_iso
        if not df_iso and not dt_iso:
            conditions.append(f"{date_expr} = (CURRENT_TIMESTAMP AT TIME ZONE :tz)::date")
        date_filter = " AND ".join(conditions)
        q = f"""
        SELECT o.id, o.operation_number, o.amount, o.payment_type_code,
               o.cashier_login, o.expeditor_login, o.customer_id, o.order_id,
               o.operation_date, o.comment, o.related_operation_id
        FROM "Sales".operations o
        WHERE {date_filter}
        ORDER BY o.operation_date DESC
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
        cols = [c for c in r.keys()]
        data = [dict(zip(cols, row)) for row in rows]
        total = sum(float(row.amount or 0) for row in rows)
        for d in data:
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()
        return {"success": True, "data": data, "total_amount": total}
    except Exception as e:
        return {"success": False, "error": str(e)[:200], "data": [], "total_amount": 0}


@router.get("/cash-received/export")
async def export_cash_received_excel(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    tz: str = Query("Asia/Tashkent"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Выгрузка принятых денег за период в Excel."""
    try:
        tz_safe = tz.replace("'", "").replace(";", "")[:50] if tz else "UTC"
        params = {"tz": tz_safe}
        date_expr = "(o.operation_date AT TIME ZONE :tz)::date"
        conditions = ["o.type_code = 'cash_receipt'", "o.status = 'completed'"]
        df_iso = _parse_date_to_iso(date_from)
        dt_iso = _parse_date_to_iso(date_to)
        if df_iso:
            conditions.append(f"{date_expr} >= :date_from")
            params["date_from"] = df_iso
        if dt_iso:
            conditions.append(f"{date_expr} <= :date_to")
            params["date_to"] = dt_iso
        if not df_iso and not dt_iso:
            conditions.append(f"{date_expr} = (CURRENT_TIMESTAMP AT TIME ZONE :tz)::date")
        q = f"""
        SELECT o.operation_number, o.amount, o.payment_type_code, o.cashier_login,
               o.expeditor_login, o.customer_id, o.order_id, o.operation_date
        FROM "Sales".operations o WHERE """ + " AND ".join(conditions) + """
        ORDER BY o.operation_date DESC
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
    except Exception:
        rows = []
    wb = Workbook()
    ws = wb.active
    ws.title = "Принятые деньги"
    headers = ["№ операции", "Сумма", "Тип оплаты", "Кассир", "От экспедитора", "Клиент ID", "Заказ №", "Дата"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows[:50000], start=2):
        for col_idx, val in enumerate(row, start=1):
            v = val.isoformat() if hasattr(val, "isoformat") else val
            ws.cell(row=row_idx, column=col_idx, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cash_received.xlsx"'},
    )


@router.get("/orders-for-confirmation/export")
async def export_orders_for_confirmation_excel(
    payment_type_code: str | None = Query(None),
    status_code: str | None = Query(None),
    payment_confirmed: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Выгрузка заказов для подтверждения оплаты в Excel."""
    try:
        conditions = [
            "o.payment_type_code IS NOT NULL",
            "o.payment_type_code != ''",
            "(o.status_code IS NULL OR o.status_code NOT IN ('cancelled', 'canceled'))",
        ]
        params = {}
        if payment_type_code and payment_type_code.strip():
            conditions.append("o.payment_type_code = :payment_type_code")
            params["payment_type_code"] = payment_type_code.strip()
        if status_code and status_code.strip():
            conditions.append("o.status_code = :status_code")
            params["status_code"] = status_code.strip()
        pc_val = (payment_confirmed or "").strip().lower()
        if pc_val == "true":
            conditions.append("EXISTS (SELECT 1 FROM \"Sales\".operations op WHERE op.type_code = 'cash_receipt' AND op.status = 'completed' AND op.order_id = o.order_no)")
        elif pc_val == "false":
            conditions.append("NOT EXISTS (SELECT 1 FROM \"Sales\".operations op WHERE op.type_code = 'cash_receipt' AND op.status = 'completed' AND op.order_id = o.order_no)")
        where_clause = " AND ".join(conditions)
        q = f"""
        SELECT o.order_no, COALESCE(c.name_client, c.firm_name, '') AS customer_name,
               c.tax_id, c.account_no, ua.fio AS agent_fio, ue.fio AS expeditor_fio,
               o.total_amount, pt.name AS payment_type_name, s.name AS status_name,
               EXISTS (SELECT 1 FROM "Sales".operations op WHERE op.type_code = 'cash_receipt'
                 AND op.status = 'completed' AND op.order_id = o.order_no) AS payment_confirmed
        FROM "Sales".orders o
        LEFT JOIN "Sales".customers c ON o.customer_id = c.id
        LEFT JOIN "Sales".payment_type pt ON o.payment_type_code = pt.code
        LEFT JOIN "Sales".status s ON o.status_code = s.code
        LEFT JOIN "Sales".users ua ON c.login_agent = ua.login
        LEFT JOIN "Sales".users ue ON c.login_expeditor = ue.login
        WHERE {where_clause}
        ORDER BY o.scheduled_delivery_at DESC NULLS LAST, o.order_date DESC
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
    except Exception:
        rows = []
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы для подтверждения"
    headers = ["№ заказа", "Клиент", "ИНН", "Р/С", "Агент", "Экспедитор", "Сумма", "Тип оплаты", "Статус", "Оплата подтверждена"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows[:50000], start=2):
        vals = [
            row[0],
            row[1] if len(row) > 1 else "",
            row[2] if len(row) > 2 else "",
            row[3] if len(row) > 3 else "",
            row[4] if len(row) > 4 else "",
            row[5] if len(row) > 5 else "",
            float(row[6]) if len(row) > 6 and row[6] is not None else "",
            row[7] if len(row) > 7 else "",
            row[8] if len(row) > 8 else "",
            "Да" if len(row) > 9 and row[9] else "Нет",
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="orders_for_confirmation.xlsx"'},
    )


@router.get("/orders-for-confirmation")
async def get_orders_for_cashier_confirmation(
    payment_type_code: str | None = Query(None, description="Фильтр по типу оплаты"),
    status_code: str | None = Query(None, description="Фильтр по статусу заказа"),
    payment_confirmed: str | None = Query(None, description="Фильтр: true=подтверждённые, false=не подтверждённые, пусто=все"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Заказы с типом оплаты для подтверждения кассиром (приём денег)."""
    try:
        conditions = [
            "o.payment_type_code IS NOT NULL",
            "o.payment_type_code != ''",
            "(o.status_code IS NULL OR o.status_code NOT IN ('cancelled', 'canceled'))",
        ]
        params = {}
        if payment_type_code and payment_type_code.strip():
            conditions.append("o.payment_type_code = :payment_type_code")
            params["payment_type_code"] = payment_type_code.strip()
        if status_code and status_code.strip():
            conditions.append("o.status_code = :status_code")
            params["status_code"] = status_code.strip()
        pc_val = (payment_confirmed or "").strip().lower()
        if pc_val == "true":
            conditions.append("""
              EXISTS (SELECT 1 FROM "Sales".operations op
                WHERE op.type_code = 'cash_receipt' AND op.status = 'completed' AND op.order_id = o.order_no)
            """)
        elif pc_val == "false":
            conditions.append("""
              NOT EXISTS (SELECT 1 FROM "Sales".operations op
                WHERE op.type_code = 'cash_receipt' AND op.status = 'completed' AND op.order_id = o.order_no)
            """)
        where_clause = " AND ".join(conditions)
        q = f"""
        SELECT o.order_no, o.customer_id, o.total_amount, o.payment_type_code, o.status_code,
               COALESCE(c.name_client, c.firm_name, '') AS customer_name,
               c.tax_id, c.account_no, c.login_agent, c.login_expeditor,
               pt.name AS payment_type_name,
               s.name AS status_name,
               ua.fio AS agent_fio,
               ue.fio AS expeditor_fio,
               EXISTS (
                 SELECT 1 FROM "Sales".operations op
                 WHERE op.type_code = 'cash_receipt' AND op.status = 'completed'
                   AND op.order_id = o.order_no
               ) AS payment_confirmed
        FROM "Sales".orders o
        LEFT JOIN "Sales".customers c ON o.customer_id = c.id
        LEFT JOIN "Sales".payment_type pt ON o.payment_type_code = pt.code
        LEFT JOIN "Sales".status s ON o.status_code = s.code
        LEFT JOIN "Sales".users ua ON c.login_agent = ua.login
        LEFT JOIN "Sales".users ue ON c.login_expeditor = ue.login
        WHERE {where_clause}
        ORDER BY o.scheduled_delivery_at DESC NULLS LAST, o.order_date DESC
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
        cols = [c for c in r.keys()]
        data = []
        for row in rows:
            d = dict(zip(cols, row))
            data.append({
                "order_no": d.get("order_no"),
                "customer_id": d.get("customer_id"),
                "customer_name": (d.get("customer_name") or "").strip() or None,
                "tax_id": d.get("tax_id"),
                "account_no": d.get("account_no"),
                "login_agent": d.get("login_agent"),
                "agent_name": (d.get("agent_fio") or d.get("login_agent") or "").strip() or None,
                "login_expeditor": d.get("login_expeditor"),
                "expeditor_name": (d.get("expeditor_fio") or d.get("login_expeditor") or "").strip() or None,
                "total_amount": float(d["total_amount"]) if d.get("total_amount") is not None else None,
                "payment_type_code": d.get("payment_type_code"),
                "payment_type_name": (d.get("payment_type_name") or d.get("payment_type_code") or "").strip() or None,
                "status_code": d.get("status_code"),
                "status_name": (d.get("status_name") or d.get("status_code") or "").strip() or None,
                "payment_confirmed": bool(d.get("payment_confirmed")),
            })
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)[:200], "data": []}


@router.get("/ledger")
async def get_financial_ledger(
    date_from: str | None = Query(None, description="Дата с (YYYY-MM-DD)"),
    date_to: str | None = Query(None, description="Дата по (YYYY-MM-DD)"),
    customer_id: int | None = Query(None, description="ID клиента"),
    movement_type: str | None = Query(None, description="Тип движения: ПРИХОД, РАСХОД, К ПОЛУЧЕНИЮ"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Финансовый реестр из VIEW v_financial_ledger."""
    try:
        q = 'SELECT * FROM "Sales".v_financial_ledger WHERE 1=1'
        params = {}
        if date_from:
            q += ' AND DATE(operation_date) >= :date_from'
            params["date_from"] = date_from
        if date_to:
            q += ' AND DATE(operation_date) <= :date_to'
            params["date_to"] = date_to
        if customer_id is not None:
            q += ' AND customer_id = :customer_id'
            params["customer_id"] = customer_id
        if movement_type:
            q += ' AND movement_type = :movement_type'
            params["movement_type"] = movement_type
        q += ' ORDER BY operation_date DESC'

        r = await session.execute(text(q), params)
        rows = r.fetchall()
        cols = [c for c in r.keys()]
        data = [dict(zip(cols, row)) for row in rows]
        for d in data:
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "error": str(e)[:200], "data": []}

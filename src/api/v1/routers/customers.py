"""
Клиенты (customers). Новая структура: id SERIAL PRIMARY KEY, остальные поля TEXT/VARCHAR.
"""
import csv
import io
from datetime import date, time
from decimal import Decimal
from src.api.v1.schemas.common import EntityModel
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.connection import get_db_session
from src.database.models import Customer, User, Operation, CustomerVisit
from src.core.deps import get_current_user, require_admin
from src.core.exceptions import ForbiddenError, NotFoundError
from src.core.pagination import PaginatedResponse, PaginationParams
from src.core.sql import escape_like
from src.api.v1.services.customer_service import CustomerService

router = APIRouter()

EXPORT_COLUMNS = [
    "id",
    "name_client",
    "firm_name",
    "category_client",
    "address",
    "city",
    "territory",
    "landmark",
    "phone",
    "contact_person",
    "tax_id",
    "status",
    "login_agent",
    "login_expeditor",
    "latitude",
    "longitude",
    "PINFL",
    "contract_no",
    "account_no",
    "bank",
    "MFO",
    "OKED",
    "VAT_code",
    "has_photo",
]

EXPORT_HEADERS_RU = [
    "ID",
    "Название клиента",
    "Название фирмы",
    "Категория",
    "Адрес",
    "Город",
    "Территория",
    "Ориентир",
    "Телефон",
    "Контактное лицо",
    "ИНН",
    "Статус",
    "Логин агента",
    "Логин экспедитора",
    "Широта",
    "Долгота",
    "ПИНФЛ",
    "Номер договора",
    "Номер счета",
    "Банк",
    "МФО",
    "ОКЭД",
    "Код НДС",
    "Фото",
]


def _customer_to_dict(c: Customer) -> dict:
    return {
        "id": c.id,
        "name_client": c.name_client,
        "firm_name": c.firm_name,
        "category_client": c.category_client,
        "address": c.address,
        "city_id": c.city_id,
        "city": getattr(c, 'city_name', None),  # Если есть JOIN
        "territory_id": c.territory_id,
        "territory": getattr(c, 'territory_name', None),  # Если есть JOIN
        "landmark": c.landmark,
        "phone": c.phone,
        "contact_person": c.contact_person,
        "tax_id": c.tax_id,
        "status": c.status,
        "login_agent": c.login_agent,
        "login_expeditor": c.login_expeditor,
        "latitude": float(c.latitude) if c.latitude is not None else None,
        "longitude": float(c.longitude) if c.longitude is not None else None,
        "PINFL": c.PINFL,
        "contract_no": c.contract_no,
        "account_no": c.account_no,
        "bank": c.bank,
        "MFO": c.MFO,
        "OKED": c.OKED,
        "VAT_code": c.VAT_code,
    }


class CustomerCreate(BaseModel):
    name_client: str | None = None
    firm_name: str | None = None
    category_client: str | None = None
    address: str | None = None
    city_id: int | None = None
    territory_id: int | None = None
    landmark: str | None = None
    phone: str | None = None
    contact_person: str | None = None
    tax_id: str | None = None
    status: str | None = None
    login_agent: str | None = None
    login_expeditor: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    PINFL: str | None = None
    contract_no: str | None = None
    account_no: str | None = None
    bank: str | None = None
    MFO: str | None = None
    OKED: str | None = None
    VAT_code: str | None = None


class CustomerUpdate(BaseModel):
    name_client: str | None = None
    firm_name: str | None = None
    category_client: str | None = None
    address: str | None = None
    city_id: int | None = None
    territory_id: int | None = None
    landmark: str | None = None
    phone: str | None = None
    contact_person: str | None = None
    tax_id: str | None = None
    status: str | None = None
    login_agent: str | None = None
    login_expeditor: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    PINFL: str | None = None
    contract_no: str | None = None
    account_no: str | None = None
    bank: str | None = None
    MFO: str | None = None
    OKED: str | None = None
    VAT_code: str | None = None


@router.get("/customers", response_model=PaginatedResponse[EntityModel])
async def list_customers(
    customer_id: int | None = Query(None, description="ID клиента (точное совпадение)"),
    search: str | None = Query(
        None,
        description="Объединённый поиск по названию клиента и ИНН (частичное совпадение, OR)",
    ),
    name_client: str | None = Query(None, description="Поиск по названию клиента (частичное совпадение)"),
    firm_name: str | None = Query(None, description="Поиск по названию фирмы"),
    city: str | None = Query(None),
    login_agent: str | None = Query(None, description="Фильтр по агенту (точно)"),
    login_expeditor: str | None = Query(None, description="Фильтр по экспедитору (точно)"),
    phone: str | None = Query(None, description="Поиск по телефону (частичное совпадение)"),
    tax_id: str | None = Query(None, description="Поиск по ИНН (частичное совпадение)"),
    pagination: PaginationParams = Depends(),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Список клиентов с фильтрами по параметрам поиска."""
    service = CustomerService(session)
    data, total = await service.list_customers(
        pagination,
        customer_id=customer_id,
        search=search,
        name_client=name_client,
        firm_name=firm_name,
        city=city,
        login_agent=login_agent,
        login_expeditor=login_expeditor,
        phone=phone,
        tax_id=tax_id,
    )
    return PaginatedResponse.create(data=data, total=total, pagination=pagination)
@router.get("/customers/export", response_model=None)
async def export_customers_excel(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Выгрузка всех клиентов в Excel (.xlsx), каждое поле в отдельной ячейке. Заголовки — русские, как в таблице. Только admin."""
    sql = """SELECT c.id, c.name_client, c.firm_name, c.category_client, c.address,
             COALESCE(ct.name, '') AS city, COALESCE(t.name, '') AS territory, c.landmark,
             c.phone, c.contact_person, c.tax_id, c.status, c.login_agent, c.login_expeditor,
             c.latitude, c.longitude, c.pinfl, c.contract_no, c.account_no, c.bank, c.mfo, c.oked, c.vat_code,
             CASE WHEN EXISTS (SELECT 1 FROM "Sales".customer_photo cp WHERE cp.customer_id = c.id) THEN 'Да' ELSE 'Нет' END
             FROM "Sales".customers c
             LEFT JOIN "Sales".cities ct ON c.city_id = ct.id
             LEFT JOIN "Sales".territories t ON c.territory_id = t.id
             ORDER BY c.id LIMIT 50000"""
    result = await session.execute(text(sql))
    rows = result.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    for col, name in enumerate(EXPORT_HEADERS_RU, start=1):
        ws.cell(row=1, column=col, value=name)
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, val in enumerate(r, start=1):
            if val is None:
                cell_val = ""
            elif isinstance(val, (int, float)):
                cell_val = val
            else:
                cell_val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=cell_val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="clients.xlsx"'},
    )


def _parse_float(s: str | None):
    if not s or not s.strip():
        return None
    try:
        return float(s.strip().replace(",", "."))
    except ValueError:
        return None


def _normalize_coordinate(value: float | None, *, field: str) -> float | None:
    if value is None:
        return None
    if field == "latitude" and not (-90 <= value <= 90):
        return None
    if field == "longitude" and not (-180 <= value <= 180):
        return None
    return value


@router.post("/customers/import", response_model=EntityModel | list[EntityModel])
async def import_customers_csv(
    file: UploadFile = File(..., description="CSV файл (разделитель ;), структура как в выгрузке"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Загрузка клиентов из CSV. Только admin. Формат: как у export (id;name_client;...)."""
    if not file.filename or not (file.filename.lower().endswith(".csv") or file.filename.lower().endswith(".txt")):
        raise HTTPException(status_code=400, detail="Нужен файл .csv или .txt")
    content = await file.read()
    try:
        text_content = content.decode("utf-8-sig") or content.decode("utf-8")
    except Exception:
        raise HTTPException(status_code=400, detail="Неверная кодировка файла (нужен UTF-8)")
    buf = io.StringIO(text_content)
    reader = csv.reader(buf, delimiter=";")
    rows = list(reader)
    if not rows:
        return {"message": "Файл пуст", "created": 0, "updated": 0}
    header = rows[0]
    if header and header[0] == "id" and "name_client" in (h.strip().lower() for h in header):
        data_rows = rows[1:]
    else:
        data_rows = rows
    created = 0
    updated = 0
    col_count = len(EXPORT_COLUMNS)
    for row in data_rows:
        if not row or len(row) < 2:
            continue
        values = (row + [""] * col_count)[:col_count]
        row_id = values[0].strip() if values[0] else ""
        try:
            pk = int(row_id) if row_id.isdigit() else None
        except ValueError:
            pk = None
        lat = _normalize_coordinate(_parse_float(values[14]), field="latitude")
        lon = _normalize_coordinate(_parse_float(values[15]), field="longitude")
        def _v(i):
            x = values[i] if i < len(values) else ""
            return x.strip() if x else None
        if pk:
            result = await session.execute(select(Customer).where(Customer.id == pk))
            existing = result.scalar_one_or_none()
            if existing:
                existing.name_client = _v(1) or existing.name_client
                existing.firm_name = _v(2) or existing.firm_name
                existing.category_client = _v(3)
                existing.address = _v(4)
                existing.city = _v(5)
                existing.territory = _v(6)
                existing.landmark = _v(7)
                existing.phone = _v(8)
                existing.contact_person = _v(9)
                existing.tax_id = _v(10)
                existing.status = _v(11) or existing.status
                existing.login_agent = _v(12)
                existing.login_expeditor = _v(13)
                existing.latitude = Decimal(str(lat)) if lat is not None else existing.latitude
                existing.longitude = Decimal(str(lon)) if lon is not None else existing.longitude
                existing.PINFL = _v(16)
                existing.contract_no = _v(17)
                existing.account_no = _v(18)
                existing.bank = _v(19)
                existing.MFO = _v(20)
                existing.OKED = _v(21)
                existing.VAT_code = _v(22)
                updated += 1
                continue
        c = Customer(
            name_client=_v(1),
            firm_name=_v(2),
            category_client=_v(3),
            address=_v(4),
            city=_v(5),
            territory=_v(6),
            landmark=_v(7),
            phone=_v(8),
            contact_person=_v(9),
            tax_id=_v(10),
            status=_v(11) or "Активный",
            login_agent=_v(12),
            login_expeditor=_v(13),
            latitude=Decimal(str(lat)) if lat is not None else None,
            longitude=Decimal(str(lon)) if lon is not None else None,
            PINFL=_v(16),
            contract_no=_v(17),
            account_no=_v(18),
            bank=_v(19),
            MFO=_v(20),
            OKED=_v(21),
            VAT_code=_v(22),
        )
        session.add(c)
        created += 1
    await session.commit()
    return {"message": "Импорт выполнен", "created": created, "updated": updated}


@router.post("/customers", response_model=EntityModel | list[EntityModel])
async def create_customer(
    body: CustomerCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Создать клиента. Доступно ролям Admin и Agent."""
    return await CustomerService(session).create_customer(body.model_dump(), user.role)
@router.get("/customers/{customer_id}/visits", response_model=EntityModel | list[EntityModel])
async def list_customer_visits(
    customer_id: int,
    limit: int = Query(100, le=500),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Список визитов клиента."""
    return await CustomerService(session).list_customer_visits(customer_id=customer_id, limit=limit)


class VisitCreateBody(BaseModel):
    visit_date: str
    visit_time: str | None = None
    status: str = "planned"
    responsible_login: str | None = None
    comment: str | None = None
@router.post("/customers/{customer_id}/visits", response_model=EntityModel | list[EntityModel])
async def create_customer_visit(
    customer_id: int,
    body: VisitCreateBody,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Создать визит для клиента."""
    return await CustomerService(session).create_customer_visit(
        customer_id=customer_id,
        payload=body.model_dump(),
        created_by=user.login,
    )
@router.get("/customers/{customer_id}", response_model=EntityModel | list[EntityModel])
async def get_customer(
    customer_id: int,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Получить клиента по id."""
    return await CustomerService(session).get_customer_dict(customer_id)
@router.get("/customers/{customer_id}/balance", response_model=EntityModel | list[EntityModel])
async def get_customer_balance(
    customer_id: int,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Баланс клиента: сумма незавершённых операций."""
    return await CustomerService(session).get_customer_balance(customer_id)
@router.patch("/customers/{customer_id}")
async def update_customer(
    customer_id: int,
    body: CustomerUpdate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Обновить данные клиента."""
    return await CustomerService(session).update_customer(
        customer_id=customer_id,
        payload=body.model_dump(exclude_unset=True),
        user_role=user.role,
        user_login=user.login,
    )
@router.delete("/customers/{customer_id}")
async def delete_customer(
    customer_id: int,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Удалить клиента. Только admin."""
    return await CustomerService(session).delete_customer(customer_id)

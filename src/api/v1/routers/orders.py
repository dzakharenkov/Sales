"""
Заказы и позиции заказа (orders, items). PK заказа — order_no (integer).
"""
import io
from datetime import datetime, timezone
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from pydantic import BaseModel
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.connection import get_db_session
from src.database.models import Order, Item, Customer, Product, Status, PaymentType, User as UserModel
from src.core.deps import get_current_user, require_admin
from src.database.models import User

router = APIRouter()


def _now_utc():
    return datetime.now(timezone.utc)


def _parse_optional_datetime(s: str | None):
    """Parse ISO date/datetime string to timezone-aware datetime or None."""
    if not s or not s.strip():
        return None
    raw = s.strip().replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(raw)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except (ValueError, TypeError):
        return None


class OrderCreate(BaseModel):
    customer_id: int | None = None
    status_code: str | None = "open"
    payment_type_code: str | None = None
    scheduled_delivery_at: str | None = None  # ISO datetime или дата


class OrderUpdate(BaseModel):
    customer_id: int | None = None
    status_code: str | None = None
    total_amount: float | None = None
    payment_type_code: str | None = None
    scheduled_delivery_at: str | None = None  # ISO datetime или дата


class ItemCreate(BaseModel):
    product_code: str
    quantity: int
    price: float | None = None


class ItemUpdate(BaseModel):
    quantity: int | None = None
    price: float | None = None


@router.get("/orders/statuses")
async def list_order_statuses(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Список статусов заказа (для выбора при создании/редактировании)."""
    result = await session.execute(select(Status.code, Status.name).order_by(Status.code))
    rows = result.all()
    return [{"code": r[0], "name": (r[1] or r[0])} for r in rows]


@router.get("/orders")
async def list_orders(
    customer_id: int | None = Query(None, description="ID клиента"),
    customer_name: str | None = Query(None, description="Поиск по названию клиента или фирмы"),
    status_code: str | None = Query(None, description="Статус заказа"),
    scheduled_delivery_from: str | None = Query(None, description="Дата поставки с (ISO дата)"),
    scheduled_delivery_to: str | None = Query(None, description="Дата поставки по (ISO дата)"),
    login_agent: str | None = Query(None, description="Логин агента"),
    login_expeditor: str | None = Query(None, description="Логин экспедитора"),
    last_updated_by: str | None = Query(None, description="Логин пользователя, выполнившего последнее изменение"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Список заказов с фильтрами: клиент, статус, дата поставки, агент, экспедитор, кто изменил. Сортировка по дате поставки."""
    q = (
        select(Order, Customer, Status, PaymentType)
        .outerjoin(Customer, Order.customer_id == Customer.id)
        .outerjoin(Status, Order.status_code == Status.code)
        .outerjoin(PaymentType, Order.payment_type_code == PaymentType.code)
        .order_by(Order.scheduled_delivery_at.desc().nullslast(), Order.order_date.desc())
    )
    if customer_id is not None:
        q = q.where(Order.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        q = q.where(or_(
            Customer.name_client.ilike(f"%{name}%"),
            Customer.firm_name.ilike(f"%{name}%"),
        ))
    if status_code and status_code.strip():
        q = q.where(Order.status_code == status_code.strip())
    if scheduled_delivery_from:
        dt_from = _parse_optional_datetime(scheduled_delivery_from)
        if dt_from:
            q = q.where(Order.scheduled_delivery_at >= dt_from)
    if scheduled_delivery_to:
        dt_to = _parse_optional_datetime(scheduled_delivery_to)
        if dt_to:
            q = q.where(Order.scheduled_delivery_at <= dt_to)
    if login_agent and login_agent.strip():
        q = q.where(Customer.login_agent == login_agent.strip())
    if login_expeditor and login_expeditor.strip():
        q = q.where(Customer.login_expeditor == login_expeditor.strip())
    if last_updated_by and last_updated_by.strip():
        q = q.where(Order.last_updated_by == last_updated_by.strip())
    
    # Подсчет общей суммы заказов (до получения данных)
    sum_q = (
        select(func.sum(func.coalesce(Order.total_amount, 0)))
        .select_from(Order)
        .outerjoin(Customer, Order.customer_id == Customer.id)
    )
    # Применяем те же фильтры
    if customer_id is not None:
        sum_q = sum_q.where(Order.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        sum_q = sum_q.where(or_(
            Customer.name_client.ilike(f"%{name}%"),
            Customer.firm_name.ilike(f"%{name}%"),
        ))
    if status_code and status_code.strip():
        sum_q = sum_q.where(Order.status_code == status_code.strip())
    if scheduled_delivery_from:
        dt_from = _parse_optional_datetime(scheduled_delivery_from)
        if dt_from:
            sum_q = sum_q.where(Order.scheduled_delivery_at >= dt_from)
    if scheduled_delivery_to:
        dt_to = _parse_optional_datetime(scheduled_delivery_to)
        if dt_to:
            sum_q = sum_q.where(Order.scheduled_delivery_at <= dt_to)
    if login_agent and login_agent.strip():
        sum_q = sum_q.where(Customer.login_agent == login_agent.strip())
    if login_expeditor and login_expeditor.strip():
        sum_q = sum_q.where(Customer.login_expeditor == login_expeditor.strip())
    if last_updated_by and last_updated_by.strip():
        sum_q = sum_q.where(Order.last_updated_by == last_updated_by.strip())
    
    total_amount_result = await session.execute(sum_q)
    total_amount_all = float(total_amount_result.scalar() or 0)
    
    result = await session.execute(q)
    rows = result.all()
    out = []
    for o, cust, st, pt in rows:
        out.append({
            "id": o.order_no,
            "order_no": o.order_no,
            "customer_id": o.customer_id,
            "customer_name": (cust.name_client or cust.firm_name or "") if cust else None,
            "order_date": o.order_date.isoformat() if o.order_date else None,
            "status_code": o.status_code,
            "status_name": (st.name if st else None) or o.status_code,
            "payment_type_code": o.payment_type_code,
            "payment_type_name": (pt.name if pt else None) or o.payment_type_code,
            "total_amount": float(o.total_amount) if o.total_amount else None,
            "created_by": o.created_by,
            "login_agent": cust.login_agent if cust else None,
            "login_expeditor": cust.login_expeditor if cust else None,
            "scheduled_delivery_at": o.scheduled_delivery_at.isoformat() if o.scheduled_delivery_at else None,
            "status_delivery_at": o.status_delivery_at.isoformat() if o.status_delivery_at else None,
            "closed_at": o.closed_at.isoformat() if o.closed_at else None,
            "last_updated_at": o.last_updated_at.isoformat() if o.last_updated_at else None,
            "last_updated_by": o.last_updated_by,
        })
    
    return {
        "orders": out,
        "total_count": len(out),
        "total_amount": total_amount_all,
    }


ORDERS_EXPORT_HEADERS_RU = [
    "№", "Клиент", "ID клиента", "Дата создания", "Статус", "Тип оплаты", "Агент", "Экспедитор",
    "Назначенная дата поставки", "Дата перевода в доставку", "Дата закрытия",
    "Последнее изменение", "Кто изменил", "Сумма", "Кто создал",
]


ORDERS_ITEMS_EXPORT_HEADERS_RU = [
    "№ заказа",
    "Клиент",
    "ID клиента",
    "Дата создания",
    "Статус",
    "Тип оплаты",
    "Код товара",
    "Товар",
    "Количество",
    "Цена",
    "Сумма",
    "Агент",
    "Экспедитор",
    "Назначенная дата поставки",
    "Дата перевода в доставку",
    "Дата закрытия",
    "Последнее изменение заказа",
    "Кто изменил заказ",
]


@router.get("/orders/export")
async def export_orders_excel(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Выгрузка заказов в Excel (.xlsx). Заголовки — русские. Только admin."""
    q = (
        select(Order, Customer, Status, PaymentType)
        .outerjoin(Customer, Order.customer_id == Customer.id)
        .outerjoin(Status, Order.status_code == Status.code)
        .outerjoin(PaymentType, Order.payment_type_code == PaymentType.code)
        .order_by(Order.order_date.desc())
    )
    result = await session.execute(q)
    rows = result.all()[:50000]
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    for col, name in enumerate(ORDERS_EXPORT_HEADERS_RU, start=1):
        ws.cell(row=1, column=col, value=name)
    for row_idx, (o, cust, st, pt) in enumerate(rows, start=2):
        customer_name = (cust.name_client or cust.firm_name or "") if cust else ""
        order_date = o.order_date.isoformat() if o.order_date else ""
        status_name = (st.name if st else None) or o.status_code or ""
        payment_name = (pt.name if pt else None) or o.payment_type_code or ""
        scheduled = o.scheduled_delivery_at.isoformat() if o.scheduled_delivery_at else ""
        status_delivery = o.status_delivery_at.isoformat() if o.status_delivery_at else ""
        closed = o.closed_at.isoformat() if o.closed_at else ""
        last_upd = o.last_updated_at.isoformat() if o.last_updated_at else ""
        login_agent = cust.login_agent if cust else ""
        login_exp = cust.login_expeditor if cust else ""
        row_data = [
            o.order_no,
            customer_name,
            o.customer_id or "",
            order_date,
            status_name,
            payment_name,
            login_agent,
            login_exp,
            scheduled,
            status_delivery,
            closed,
            last_upd,
            o.last_updated_by or "",
            float(o.total_amount) if o.total_amount is not None else "",
            o.created_by or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="orders.xlsx"'},
    )


@router.get("/orders/items")
async def list_order_items(
    customer_id: int | None = Query(None, description="ID клиента"),
    customer_name: str | None = Query(None, description="Поиск по названию клиента или фирмы"),
    status_code: str | None = Query(None, description="Статус заказа"),
    scheduled_delivery_from: str | None = Query(None, description="Дата поставки с (ISO дата)"),
    scheduled_delivery_to: str | None = Query(None, description="Дата поставки по (ISO дата)"),
    login_agent: str | None = Query(None, description="Логин агента"),
    login_expeditor: str | None = Query(None, description="Логин экспедитора"),
    last_updated_by: str | None = Query(None, description="Логин пользователя, выполнившего последнее изменение заказа"),
    limit: int = Query(100, ge=1, le=1000, description="Количество записей на странице"),
    offset: int = Query(0, ge=0, description="Смещение для пагинации"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """
    Список позиций заказов (items) с теми же фильтрами, что и /orders.
    Каждая строка — одна позиция товара в заказе.
    Поддерживает пагинацию (limit, offset).
    """
    q = (
        select(Item, Order, Customer, Product, Status, PaymentType)
        .join(Order, Item.order_id == Order.order_no)
        .outerjoin(Customer, Order.customer_id == Customer.id)
        .outerjoin(Product, Item.product_code == Product.code)
        .outerjoin(Status, Order.status_code == Status.code)
        .outerjoin(PaymentType, Order.payment_type_code == PaymentType.code)
        .order_by(Order.scheduled_delivery_at.desc().nullslast(), Order.order_date.desc(), Item.id)
    )
    if customer_id is not None:
        q = q.where(Order.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        q = q.where(
            or_(
                Customer.name_client.ilike(f"%{name}%"),
                Customer.firm_name.ilike(f"%{name}%"),
            )
        )
    if status_code and status_code.strip():
        q = q.where(Order.status_code == status_code.strip())
    if scheduled_delivery_from:
        dt_from = _parse_optional_datetime(scheduled_delivery_from)
        if dt_from:
            q = q.where(Order.scheduled_delivery_at >= dt_from)
    if scheduled_delivery_to:
        dt_to = _parse_optional_datetime(scheduled_delivery_to)
        if dt_to:
            q = q.where(Order.scheduled_delivery_at <= dt_to)
    if login_agent and login_agent.strip():
        q = q.where(Customer.login_agent == login_agent.strip())
    if login_expeditor and login_expeditor.strip():
        q = q.where(Customer.login_expeditor == login_expeditor.strip())
    if last_updated_by and last_updated_by.strip():
        q = q.where(Order.last_updated_by == last_updated_by.strip())

    # Подсчет общего количества (до пагинации)
    from sqlalchemy import func
    count_q = select(func.count()).select_from(q.subquery())
    total_count_result = await session.execute(count_q)
    total_count = total_count_result.scalar() or 0

    # Применяем пагинацию
    q = q.limit(limit).offset(offset)

    result = await session.execute(q)
    rows = result.all()
    out: list[dict] = []
    total_amount = 0.0
    for it, o, cust, prod, st, pt in rows:
        qty = it.quantity or 0
        price = float(it.price) if it.price is not None else None
        amount = qty * (price or 0)
        total_amount += amount
        out.append(
            {
                "item_id": str(it.id),
                "order_no": o.order_no,
                "customer_id": o.customer_id,
                "customer_name": (cust.name_client or cust.firm_name or "") if cust else None,
                "order_date": o.order_date.isoformat() if o.order_date else None,
                "status_code": o.status_code,
                "status_name": (st.name if st else None) or o.status_code,
                "payment_type_code": o.payment_type_code,
                "payment_type_name": (pt.name if pt else None) or o.payment_type_code,
                "product_code": it.product_code,
                "product_name": prod.name if prod else None,
                "quantity": qty,
                "price": price,
                "amount": amount,
                "login_agent": cust.login_agent if cust else None,
                "login_expeditor": cust.login_expeditor if cust else None,
                "scheduled_delivery_at": o.scheduled_delivery_at.isoformat() if o.scheduled_delivery_at else None,
                "status_delivery_at": o.status_delivery_at.isoformat() if o.status_delivery_at else None,
                "closed_at": o.closed_at.isoformat() if o.closed_at else None,
                "order_last_updated_at": o.last_updated_at.isoformat() if o.last_updated_at else None,
                "order_last_updated_by": o.last_updated_by,
            }
        )
    
    # Подсчет общей суммы всех позиций (не только текущей страницы)
    # Для этого нужно выполнить отдельный запрос
    sum_q = (
        select(func.sum(Item.quantity * func.coalesce(Item.price, 0)))
        .select_from(Item)
        .join(Order, Item.order_id == Order.order_no)
        .outerjoin(Customer, Order.customer_id == Customer.id)
    )
    # Применяем те же фильтры
    if customer_id is not None:
        sum_q = sum_q.where(Order.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        sum_q = sum_q.where(
            or_(
                Customer.name_client.ilike(f"%{name}%"),
                Customer.firm_name.ilike(f"%{name}%"),
            )
        )
    if status_code and status_code.strip():
        sum_q = sum_q.where(Order.status_code == status_code.strip())
    if scheduled_delivery_from:
        dt_from = _parse_optional_datetime(scheduled_delivery_from)
        if dt_from:
            sum_q = sum_q.where(Order.scheduled_delivery_at >= dt_from)
    if scheduled_delivery_to:
        dt_to = _parse_optional_datetime(scheduled_delivery_to)
        if dt_to:
            sum_q = sum_q.where(Order.scheduled_delivery_at <= dt_to)
    if login_agent and login_agent.strip():
        sum_q = sum_q.where(Customer.login_agent == login_agent.strip())
    if login_expeditor and login_expeditor.strip():
        sum_q = sum_q.where(Customer.login_expeditor == login_expeditor.strip())
    if last_updated_by and last_updated_by.strip():
        sum_q = sum_q.where(Order.last_updated_by == last_updated_by.strip())
    
    total_amount_result = await session.execute(sum_q)
    total_amount_all = float(total_amount_result.scalar() or 0)
    
    return {
        "items": out,
        "total_count": total_count,
        "total_amount": total_amount_all,
        "limit": limit,
        "offset": offset,
    }


@router.get("/orders/items/export")
async def export_order_items_excel(
    customer_id: int | None = Query(None, description="ID клиента"),
    customer_name: str | None = Query(None, description="Поиск по названию клиента или фирмы"),
    status_code: str | None = Query(None, description="Статус заказа"),
    scheduled_delivery_from: str | None = Query(None, description="Дата поставки с (ISO дата)"),
    scheduled_delivery_to: str | None = Query(None, description="Дата поставки по (ISO дата)"),
    login_agent: str | None = Query(None, description="Логин агента"),
    login_expeditor: str | None = Query(None, description="Логин экспедитора"),
    last_updated_by: str | None = Query(None, description="Логин пользователя, выполнившего последнее изменение заказа"),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(require_admin),
):
    """Выгрузка позиций заказов в Excel (.xlsx). Только admin."""
    q = (
        select(Item, Order, Customer, Product, Status, PaymentType)
        .join(Order, Item.order_id == Order.order_no)
        .outerjoin(Customer, Order.customer_id == Customer.id)
        .outerjoin(Product, Item.product_code == Product.code)
        .outerjoin(Status, Order.status_code == Status.code)
        .outerjoin(PaymentType, Order.payment_type_code == PaymentType.code)
        .order_by(Order.order_date.desc(), Item.id)
    )
    if customer_id is not None:
        q = q.where(Order.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        q = q.where(
            or_(
                Customer.name_client.ilike(f"%{name}%"),
                Customer.firm_name.ilike(f"%{name}%"),
            )
        )
    if status_code and status_code.strip():
        q = q.where(Order.status_code == status_code.strip())
    if scheduled_delivery_from:
        dt_from = _parse_optional_datetime(scheduled_delivery_from)
        if dt_from:
            q = q.where(Order.scheduled_delivery_at >= dt_from)
    if scheduled_delivery_to:
        dt_to = _parse_optional_datetime(scheduled_delivery_to)
        if dt_to:
            q = q.where(Order.scheduled_delivery_at <= dt_to)
    if login_agent and login_agent.strip():
        q = q.where(Customer.login_agent == login_agent.strip())
    if login_expeditor and login_expeditor.strip():
        q = q.where(Customer.login_expeditor == login_expeditor.strip())
    if last_updated_by and last_updated_by.strip():
        q = q.where(Order.last_updated_by == last_updated_by.strip())

    result = await session.execute(q)
    rows = result.all()[:50000]
    wb = Workbook()
    ws = wb.active
    ws.title = "Позиции заказов"
    for col, name in enumerate(ORDERS_ITEMS_EXPORT_HEADERS_RU, start=1):
        ws.cell(row=1, column=col, value=name)
    for row_idx, (it, o, cust, prod, st, pt) in enumerate(rows, start=2):
        customer_name_val = (cust.name_client or cust.firm_name or "") if cust else ""
        order_date = o.order_date.isoformat() if o.order_date else ""
        status_name = (st.name if st else None) or o.status_code or ""
        payment_name = (pt.name if pt else None) or o.payment_type_code or ""
        qty = it.quantity or 0
        price = float(it.price) if it.price is not None else None
        amount = qty * (price or 0)
        scheduled = o.scheduled_delivery_at.isoformat() if o.scheduled_delivery_at else ""
        status_delivery = o.status_delivery_at.isoformat() if o.status_delivery_at else ""
        closed = o.closed_at.isoformat() if o.closed_at else ""
        last_upd = o.last_updated_at.isoformat() if o.last_updated_at else ""
        row_data = [
            o.order_no,
            customer_name_val,
            o.customer_id or "",
            order_date,
            status_name,
            payment_name,
            it.product_code or "",
            prod.name if prod else "",
            qty,
            price,
            amount,
            cust.login_agent if cust else "",
            cust.login_expeditor if cust else "",
            scheduled,
            status_delivery,
            closed,
            last_upd,
            o.last_updated_by or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val if val not in ("", None) else None)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="order_items.xlsx"'},
    )


@router.post("/orders")
async def create_order(
    body: OrderCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Создать заказ. Позиции сохраняются в таблицу Sales.items (POST /orders/{id}/items)."""
    try:
        next_no = await session.execute(select(func.coalesce(func.max(Order.order_no), 0) + 1))
        order_no = next_no.scalar() or 1
        order = Order(
            order_no=order_no,
            customer_id=body.customer_id,
            status_code=body.status_code or "open",
            payment_type_code=body.payment_type_code,
            created_by=user.login,
            scheduled_delivery_at=_parse_optional_datetime(body.scheduled_delivery_at),
        )
        session.add(order)
        await session.commit()
        await session.refresh(order)
        return {"id": order.order_no, "order_no": order.order_no, "status_code": order.status_code, "message": "created"}
    except HTTPException:
        raise
    except Exception as e:
        msg = str(e)
        # Сообщение «колонка отсутствует» только если в ошибке явно указано, что колонки нет
        if "payment_type_code" in msg and "does not exist" in msg:
            raise HTTPException(
                status_code=500,
                detail="В таблице orders отсутствует колонка payment_type_code. Выполните в БД: ALTER TABLE \"Sales\".orders ADD COLUMN payment_type_code VARCHAR REFERENCES \"Sales\".payment_type(code);",
            )
        raise HTTPException(status_code=500, detail="Ошибка при создании заказа: " + msg)


@router.get("/orders/{order_id}")
async def get_order(
    order_id: int,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Заказ по номеру (order_no), позиции, клиент, агент, экспедитор, тип оплаты."""
    result = await session.execute(select(Order).where(Order.order_no == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    cust_result = await session.execute(select(Customer).where(Customer.id == order.customer_id)) if order.customer_id else None
    cust = cust_result.scalar_one_or_none() if cust_result else None
    agent_fio = expeditor_fio = None
    if cust:
        if cust.login_agent:
            u = await session.execute(select(UserModel).where(UserModel.login == cust.login_agent))
            agent_user = u.scalar_one_or_none()
            agent_fio = agent_user.fio if agent_user else None
        if cust.login_expeditor:
            u2 = await session.execute(select(UserModel).where(UserModel.login == cust.login_expeditor))
            exp_user = u2.scalar_one_or_none()
            expeditor_fio = exp_user.fio if exp_user else None
    items_result = await session.execute(select(Item).where(Item.order_id == order.order_no))
    items = items_result.scalars().all()
    return {
        "id": order.order_no,
        "order_no": order.order_no,
        "customer_id": order.customer_id,
        "customer_name": (cust.name_client or cust.firm_name or "") if cust else None,
        "order_date": order.order_date.isoformat() if order.order_date else None,
        "status_code": order.status_code,
        "total_amount": float(order.total_amount) if order.total_amount else None,
        "payment_type_code": order.payment_type_code,
        "created_by": order.created_by,
        "login_agent": cust.login_agent if cust else None,
        "login_expeditor": cust.login_expeditor if cust else None,
        "agent_fio": agent_fio,
        "expeditor_fio": expeditor_fio,
        "scheduled_delivery_at": order.scheduled_delivery_at.isoformat() if order.scheduled_delivery_at else None,
        "status_delivery_at": order.status_delivery_at.isoformat() if order.status_delivery_at else None,
        "closed_at": order.closed_at.isoformat() if order.closed_at else None,
        "last_updated_at": order.last_updated_at.isoformat() if order.last_updated_at else None,
        "last_updated_by": order.last_updated_by,
        "items": [
            {
                "id": str(i.id),
                "product_code": i.product_code,
                "quantity": i.quantity,
                "price": float(i.price) if i.price else None,
            }
            for i in items
        ],
    }


@router.patch("/orders/{order_id}")
async def update_order(
    order_id: int,
    body: OrderUpdate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Изменить заказ по номеру (клиент, статус, тип оплаты, сумма, назначенная дата поставки)."""
    result = await session.execute(select(Order).where(Order.order_no == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    now = _now_utc()
    if body.customer_id is not None:
        order.customer_id = body.customer_id
    if body.status_code is not None:
        order.status_code = body.status_code
        # Определяем по названию/коду статуса, проставлять ли даты (работает для любых кодов: 1, 2, 3, 4, delivered, отмена и т.д.)
        st_result = await session.execute(select(Status).where(Status.code == body.status_code))
        st_row = st_result.scalar_one_or_none()
        name_lower = ((st_row.name or "") + " " + (body.status_code or "")).strip().lower()
        if order.status_delivery_at is None and any(k in name_lower for k in ("достав", "deliver", "delivery", "shipping")):
            order.status_delivery_at = now
        if order.closed_at is None and any(k in name_lower for k in ("отмен", "cancel", "closed", "доставлен", "delivered", "dismiss")):
            order.closed_at = now
    if body.total_amount is not None:
        order.total_amount = body.total_amount
    if body.payment_type_code is not None:
        order.payment_type_code = body.payment_type_code
    if body.scheduled_delivery_at is not None:
        order.scheduled_delivery_at = _parse_optional_datetime(body.scheduled_delivery_at)
    order.last_updated_at = now
    order.last_updated_by = user.login
    await session.commit()
    await session.refresh(order)
    return {"id": order.order_no, "message": "updated"}


@router.post("/orders/{order_id}/items")
async def add_order_item(
    order_id: int,
    body: ItemCreate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Добавить позицию в заказ (order_id = номер заказа)."""
    result = await session.execute(select(Order).where(Order.order_no == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    item = Item(
        order_id=order_id,
        product_code=body.product_code,
        quantity=body.quantity,
        price=body.price,
        last_updated_by=user.login,
    )
    session.add(item)
    await session.commit()
    await session.refresh(item)
    return {"id": str(item.id), "product_code": item.product_code, "quantity": item.quantity, "message": "created"}


@router.patch("/orders/{order_id}/items/{item_id}")
async def update_order_item(
    order_id: int,
    item_id: UUID,
    body: ItemUpdate,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Изменить позицию заказа."""
    result = await session.execute(select(Item).where(Item.id == item_id).where(Item.order_id == order_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    if body.quantity is not None:
        item.quantity = body.quantity
    if body.price is not None:
        item.price = body.price
    item.last_updated_by = user.login
    await session.commit()
    await session.refresh(item)
    return {"id": str(item.id), "message": "updated"}


@router.delete("/orders/{order_id}/items/{item_id}")
async def delete_order_item(
    order_id: int,
    item_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Удалить позицию из заказа."""
    result = await session.execute(select(Item).where(Item.id == item_id).where(Item.order_id == order_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    await session.delete(item)
    await session.commit()
    return {"message": "deleted"}

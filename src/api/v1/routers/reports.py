"""
Отчётность: по клиентам, агентам, экспедиторам, визитам, дашборд, фото.
"""
import io
from datetime import date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.connection import get_db_session
from src.core.deps import get_current_user
from src.database.models import User

router = APIRouter()


def _parse_month(s: str | None) -> str | None:
    """Приводит месяц к YYYY-MM. Поддерживает YYYY-MM, MM.YYYY, MM/YYYY."""
    if not s or not (s := str(s).strip()):
        return None
    s = s[:7]
    if "-" in s:
        parts = s.split("-")
        if len(parts) == 2:
            try:
                y, m = int(parts[0]), int(parts[1])
                if 1 <= m <= 12 and 1900 <= y <= 2100:
                    return f"{y:04d}-{m:02d}"
            except (ValueError, TypeError):
                pass
    if "." in s:
        parts = s.split(".")
        if len(parts) == 2:
            try:
                m, y = int(parts[0]), int(parts[1])
                if 1 <= m <= 12 and 1900 <= y <= 2100:
                    return f"{y:04d}-{m:02d}"
            except (ValueError, TypeError):
                pass
    return None


def _parse_date_to_iso(s: str | None) -> str | None:
    """Приводит дату к YYYY-MM-DD. Поддерживает Y-m-d и d.m.Y."""
    if not s or not (s := str(s).strip()):
        return None
    s = s[:10]
    if len(s) == 10:
        if "-" in s and s[4] == "-" and s[7] == "-":
            return s
        if "." in s:
            parts = s.split(".")
            if len(parts) == 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if 1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100:
                        return f"{y:04d}-{m:02d}-{d:02d}"
                except (ValueError, TypeError):
                    pass
    return None


def _iso_to_date(s: str | None) -> date | None:
    """Преобразует ISO-дату YYYY-MM-DD в объект date для asyncpg."""
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


@router.get("/customers")
async def report_customers(
    status: str | None = Query(None),
    agent_login: str | None = Query(None),
    month: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Отчёт по клиентам: визиты, заказы."""
    try:
        month_iso = _parse_month(month)
        df = _parse_date_to_iso(date_from)
        dt = _parse_date_to_iso(date_to)
        params = {}
        conditions = []
        if status and status != "all":
            if status == "active":
                conditions.append("rc.status = 'Активен'")
            elif status == "inactive":
                conditions.append("rc.status = 'Неактивен'")
        if agent_login:
            conditions.append("rc.login_agent = :agent_login")
            params["agent_login"] = agent_login
        if month_iso:
            conditions.append("(rc.last_visit_date >= :m_start AND rc.last_visit_date < :m_end)")
            params["m_start"] = _iso_to_date(month_iso + "-01") or (month_iso + "-01")
            y, m = int(month_iso[:4]), int(month_iso[5:7])
            end_str = f"{y}-{m+1:02d}-01" if m < 12 else f"{y+1}-01-01"
            params["m_end"] = _iso_to_date(end_str) or end_str
        if df and dt:
            conditions.append("rc.last_visit_date >= :date_from AND rc.last_visit_date <= :date_to")
            params["date_from"] = _iso_to_date(df) or df
            params["date_to"] = _iso_to_date(dt) or dt
        where_sql = " AND ".join(conditions) if conditions else "1=1"
        q = f"""
        SELECT rc.id, rc.name_client, rc.firm_name, rc.login_agent, c.login_expeditor,
               rc.total_visits, rc.completed_visits, rc.last_visit_date,
               ua.fio AS agent_fio, ue.fio AS expeditor_fio,
               COALESCE(o.cnt, 0)::int AS orders_count, COALESCE(o.amt, 0) AS orders_amount
        FROM "Sales".v_report_customers rc
        JOIN "Sales".customers c ON c.id = rc.id
        LEFT JOIN "Sales".users ua ON rc.login_agent = ua.login
        LEFT JOIN "Sales".users ue ON c.login_expeditor = ue.login
        LEFT JOIN (
          SELECT customer_id, COUNT(*)::int AS cnt, SUM(total_amount) AS amt
          FROM "Sales".orders
          WHERE status_code IS NOT NULL AND status_code NOT IN ('cancelled', 'canceled')
          GROUP BY customer_id
        ) o ON rc.id = o.customer_id
        WHERE {where_sql}
        ORDER BY rc.last_visit_date DESC NULLS LAST
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
        cols = list(r.keys())
        data = []
        for row in rows:
            d = dict(zip(cols, row))
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()
            data.append(d)
        return {"total": len(data), "data": data}
    except Exception as e:
        return {"total": 0, "data": [], "error": str(e)[:200]}


@router.get("/customers/export")
async def report_customers_export(
    status: str | None = Query(None),
    agent_login: str | None = Query(None),
    month: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Экспорт отчёта по клиентам в Excel."""
    res = await report_customers(status, agent_login, month, date_from, date_to, session, user)
    data = res.get("data") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "По клиентам"
    def _client_display(r):
        n = (r.get("name_client") or "").strip()
        f = (r.get("firm_name") or "").strip()
        if f and f != n:
            return f"{n} ({f})" if n else f
        return n or f or ""
    def _visit_completed(r):
        return "Визит завершён" if (r.get("completed_visits") or 0) > 0 else "—"
    headers = ["Клиент", "ФИО Агента", "ФИО Экспедитора", "Визит завершён", "Кол-во заказов", "Сумма заказов"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(data[:50000], start=2):
        vals = [
            _client_display(r),
            r.get("agent_fio") or "",
            r.get("expeditor_fio") or "",
            _visit_completed(r),
            r.get("orders_count") or 0,
            float(r.get("orders_amount") or 0),
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_customers.xlsx"'},
    )


@router.get("/agents")
async def report_agents(
    month: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Отчёт по агентам."""
    try:
        month_iso = _parse_month(month)
        df = _parse_date_to_iso(date_from)
        dt = _parse_date_to_iso(date_to)
        params = {}
        date_filter = ""
        if month_iso:
            date_filter = " AND cv.visit_date >= :m_start AND cv.visit_date < :m_end"
            params["m_start"] = _iso_to_date(month_iso + "-01") or (month_iso + "-01")
            y, m = int(month_iso[:4]), int(month_iso[5:7])
            end_str = f"{y}-{m+1:02d}-01" if m < 12 else f"{y+1}-01-01"
            params["m_end"] = _iso_to_date(end_str) or end_str
        elif df or dt:
            if df and dt:
                date_filter = " AND cv.visit_date >= :date_from AND cv.visit_date <= :date_to"
                params["date_from"] = _iso_to_date(df) or df
                params["date_to"] = _iso_to_date(dt) or dt
            elif df:
                date_filter = " AND cv.visit_date >= :date_from"
                params["date_from"] = _iso_to_date(df) or df
            else:
                date_filter = " AND cv.visit_date <= :date_to"
                params["date_to"] = _iso_to_date(dt) or dt
        date_filter_orders = date_filter.replace("cv.visit_date", "o.order_date::date")
        q = f"""
        SELECT u.login, u.fio,
               (SELECT COUNT(*)::int FROM "Sales".customers c WHERE c.login_agent = u.login) AS client_count,
               COUNT(DISTINCT cv.id) AS total_visits,
               SUM(CASE WHEN cv.status = 'completed' THEN 1 ELSE 0 END)::int AS completed_visits,
               ROUND(SUM(CASE WHEN cv.status = 'completed' THEN 1 ELSE 0 END)::numeric / NULLIF(COUNT(DISTINCT cv.id), 0) * 100, 1) AS visit_completion_rate,
               COALESCE((SELECT SUM(o.total_amount) FROM "Sales".orders o
                 JOIN "Sales".customers c ON o.customer_id = c.id AND c.login_agent = u.login
                 WHERE o.status_code IS NOT NULL AND o.status_code NOT IN ('canceled', 'cancelled'){date_filter_orders}), 0) AS orders_amount,
               COALESCE((SELECT COUNT(*)::int FROM "Sales".orders o
                 JOIN "Sales".customers c ON o.customer_id = c.id AND c.login_agent = u.login
                 WHERE o.status_code IS NOT NULL AND o.status_code NOT IN ('canceled', 'cancelled'){date_filter_orders}), 0) AS orders_count,
               COALESCE((SELECT COUNT(*)::int FROM "Sales".orders o
                 JOIN "Sales".customers c ON o.customer_id = c.id AND c.login_agent = u.login
                 WHERE o.status_code = 'completed'{date_filter_orders}), 0) AS orders_completed
        FROM "Sales".users u
        LEFT JOIN "Sales".customers_visits cv ON u.login = cv.responsible_login {date_filter}
        WHERE LOWER(u.role::text) IN ('agent', 'expeditor', 'admin')
        GROUP BY u.login, u.fio
        ORDER BY total_visits DESC
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
        cols = list(r.keys())
        data = [dict(zip(cols, row)) for row in rows]
        for d in data:
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()
            # Процент завершённых заказов
            oc = d.get("orders_count") or 0
            ocomp = d.get("orders_completed") or 0
            d["orders_completion_rate"] = round(ocomp / oc * 100, 1) if oc else 0
        return {"data": data}
    except Exception as e:
        return {"data": [], "error": str(e)[:200]}


@router.get("/agents/export")
async def report_agents_export(
    month: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Экспорт отчёта по агентам в Excel."""
    res = await report_agents(month, date_from, date_to, session, user)
    data = res.get("data") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Агенты"
    headers = ["Логин", "ФИО", "Клиентов", "Визитов", "Завершено", "% завершённости визитов", "Сумма заказов", "Кол-во заказов", "% завершённости заказов"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(data[:50000], start=2):
        visit_rate = r.get("visit_completion_rate") or 0
        order_rate = r.get("orders_completion_rate") or 0
        vals = [
            r.get("login") or "",
            r.get("fio") or "",
            r.get("client_count") or 0,
            r.get("total_visits") or 0,
            r.get("completed_visits") or 0,
            visit_rate,
            float(r.get("orders_amount") or 0),
            r.get("orders_count") or 0,
            order_rate,
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_agents.xlsx"'},
    )


@router.get("/expeditors")
async def report_expeditors(
    month: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Сводный отчёт по экспедиторам на основе заказов."""
    try:
        month_iso = _parse_month(month)
        df = _parse_date_to_iso(date_from)
        dt = _parse_date_to_iso(date_to)
        params = {}
        date_filter = ""
        if month_iso:
            date_filter = " AND (o.order_date::date >= :m_start AND o.order_date::date < :m_end)"
            params["m_start"] = _iso_to_date(month_iso + "-01") or (month_iso + "-01")
            y, m = int(month_iso[:4]), int(month_iso[5:7])
            end_str = f"{y}-{m+1:02d}-01" if m < 12 else f"{y+1}-01-01"
            params["m_end"] = _iso_to_date(end_str) or end_str
        elif df or dt:
            if df and dt:
                date_filter = " AND (o.order_date::date >= :date_from AND o.order_date::date <= :date_to)"
                params["date_from"] = _iso_to_date(df) or df
                params["date_to"] = _iso_to_date(dt) or dt
            elif df:
                date_filter = " AND o.order_date::date >= :date_from"
                params["date_from"] = _iso_to_date(df) or df
            else:
                date_filter = " AND o.order_date::date <= :date_to"
                params["date_to"] = _iso_to_date(dt) or dt
        q = f"""
        SELECT u.login, u.fio,
               COUNT(DISTINCT o.order_no)::int AS orders_count,
               COALESCE(SUM(o.total_amount), 0) AS orders_amount,
               SUM(CASE WHEN o.status_code = 'open' THEN 1 ELSE 0 END)::int AS orders_open,
               SUM(CASE WHEN o.status_code = 'delivery' THEN 1 ELSE 0 END)::int AS orders_delivery,
               SUM(CASE WHEN o.status_code = 'completed' THEN 1 ELSE 0 END)::int AS orders_completed,
               SUM(CASE WHEN o.status_code IN ('canceled', 'cancelled') THEN 1 ELSE 0 END)::int AS orders_cancelled
        FROM "Sales".users u
        LEFT JOIN "Sales".customers c ON c.login_expeditor = u.login
        LEFT JOIN "Sales".orders o ON o.customer_id = c.id {date_filter}
        WHERE LOWER(u.role::text) = 'expeditor'
        GROUP BY u.login, u.fio
        ORDER BY orders_count DESC NULLS LAST
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
        cols = list(r.keys())
        data = []
        for row in rows:
            d = dict(zip(cols, row))
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()
                elif hasattr(v, "__float__"):
                    d[k] = float(v) if v is not None else 0
            data.append(d)
        return {"data": data}
    except Exception as e:
        return {"data": [], "error": str(e)[:200]}


@router.get("/expeditors/export")
async def report_expeditors_export(
    month: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Экспорт сводного отчёта по экспедиторам в Excel."""
    res = await report_expeditors(month, date_from, date_to, session, user)
    data = res.get("data") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Экспедиторы"
    headers = ["Логин", "ФИО", "Заказов", "Сумма", "Открыто", "В доставке", "Доставлено", "Отменено"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(data[:50000], start=2):
        vals = [
            r.get("login") or "",
            r.get("fio") or "",
            r.get("orders_count") or 0,
            float(r.get("orders_amount") or 0),
            r.get("orders_open") or 0,
            r.get("orders_delivery") or 0,
            r.get("orders_completed") or 0,
            r.get("orders_cancelled") or 0,
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_expeditors.xlsx"'},
    )


@router.get("/visits")
async def report_visits(
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Отчёт по визитам — статистика по датам."""
    try:
        df = _parse_date_to_iso(from_date)
        dt = _parse_date_to_iso(to_date)
        params = {}
        conditions = []
        if df:
            conditions.append("DATE(cv.visit_date) >= :date_from")
            params["date_from"] = _iso_to_date(df) or df
        if dt:
            conditions.append("DATE(cv.visit_date) <= :date_to")
            params["date_to"] = _iso_to_date(dt) or dt
        where_sql = " AND ".join(conditions) if conditions else "1=1"
        q = f"""
        SELECT DATE(cv.visit_date)::text AS date,
               COUNT(*)::int AS total_visits,
               SUM(CASE WHEN cv.status = 'completed' THEN 1 ELSE 0 END)::int AS completed,
               SUM(CASE WHEN cv.status = 'planned' THEN 1 ELSE 0 END)::int AS planned,
               SUM(CASE WHEN cv.status = 'cancelled' THEN 1 ELSE 0 END)::int AS cancelled
        FROM "Sales".customers_visits cv
        WHERE {where_sql}
        GROUP BY DATE(cv.visit_date)
        ORDER BY date DESC
        """
        r = await session.execute(text(q), params)
        rows = r.fetchall()
        by_date = []
        for row in rows:
            d = dict(zip(r.keys(), row))
            for k, v in list(d.items()):
                if hasattr(v, "isoformat"):
                    d[k] = v.isoformat()[:10] if hasattr(v, "isoformat") else str(v)[:10]
            by_date.append(d)
        total_visits = sum(int(r.get("total_visits", 0) or 0) for r in by_date)
        completed = sum(int(r.get("completed", 0) or 0) for r in by_date)
        rate = round(completed / total_visits * 100, 1) if total_visits else 0
        summary = {"total_visits": total_visits, "completed": completed, "completion_rate": rate}
        return {"summary": summary, "by_date": by_date}
    except Exception as e:
        return {"summary": {}, "by_date": [], "error": str(e)[:200]}


@router.get("/visits/export")
async def report_visits_export(
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Экспорт отчёта по визитам в Excel."""
    res = await report_visits(from_date, to_date, session, user)
    by_date = res.get("by_date") or []
    summary = res.get("summary") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитика визитов"
    headers = ["Дата", "Всего", "Завершено", "Запланировано", "Отменено"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(by_date[:50000], start=2):
        vals = [
            r.get("date") or "",
            r.get("total_visits") or 0,
            r.get("completed") or 0,
            r.get("planned") or 0,
            r.get("cancelled") or 0,
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    row_final = len(by_date) + 2
    ws.cell(row=row_final, column=1, value="Итого")
    ws.cell(row=row_final, column=2, value=summary.get("total_visits") or 0)
    ws.cell(row=row_final, column=3, value=summary.get("completed") or 0)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_visits.xlsx"'},
    )


@router.get("/dashboard")
async def report_dashboard(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    status_codes: str | None = Query(None),
    product_category: str | None = Query(None),
    territory: str | None = Query(None),
    category_client: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Сводная аналитика: заказы, визиты, по категориям продуктов."""
    try:
        from datetime import date as dt_date
        df = _parse_date_to_iso(date_from)
        dt = _parse_date_to_iso(date_to)
        today_str = dt_date.today().isoformat()
        if df and dt:
            date_cond = "o.order_date::date >= :date_from AND o.order_date::date <= :date_to"
        elif df:
            date_cond = "o.order_date::date >= :date_from"
        elif dt:
            date_cond = "o.order_date::date <= :date_to"
        else:
            date_cond = "o.order_date::date = CURRENT_DATE"

        params = {}
        if df:
            d = _iso_to_date(df)
            params["date_from"] = d if d is not None else df
        if dt:
            d = _iso_to_date(dt)
            params["date_to"] = d if d is not None else dt

        status_list = []
        if status_codes and status_codes.strip():
            for s in status_codes.replace(",", " ").split():
                ss = s.strip().lower()
                if ss and ss not in ("canceled", "cancelled"):
                    status_list.append(ss)
        if not status_list:
            status_list = ["open", "delivery", "completed"]
        params["status_codes"] = status_list
        status_cond = "o.status_code = ANY(:status_codes)"

        extra_joins = ""
        extra_where = []
        if product_category and product_category.strip():
            extra_joins += ' JOIN "Sales".items it0 ON it0.order_id = o.order_no JOIN "Sales".product pr0 ON pr0.code = it0.product_code AND pr0.type_id = :product_category'
            params["product_category"] = product_category.strip()
        if territory and territory.strip():
            extra_where.append("c.territory = :territory")
            params["territory"] = territory.strip()
        if category_client and category_client.strip():
            extra_where.append("c.category_client = :category_client")
            params["category_client"] = category_client.strip()
        cust_join = 'LEFT JOIN "Sales".customers c ON o.customer_id = c.id'
        extra_where_sql = " AND " + " AND ".join(extra_where) if extra_where else ""

        q_sum = f"""
        SELECT COALESCE(SUM(o.total_amount), 0) AS total_sum
        FROM "Sales".orders o
        {cust_join} {extra_joins}
        WHERE {date_cond} AND {status_cond} {extra_where_sql}
        """
        r_sum = await session.execute(text(q_sum), params)
        total_sum = float(r_sum.scalar() or 0)

        q_cat = f"""
        SELECT COALESCE(pt.name, 'Без категории') AS category_code,
               COALESCE(NULLIF(TRIM(pt.description), ''), pt.name, 'Без категории') AS category_name,
               COALESCE(SUM(it.quantity * COALESCE(it.price, 0)), 0) AS sum_amount,
               COALESCE(SUM(it.quantity), 0)::int AS total_quantity
        FROM "Sales".orders o
        {cust_join}
        JOIN "Sales".items it ON it.order_id = o.order_no
        JOIN "Sales".product p ON p.code = it.product_code
        LEFT JOIN "Sales".product_type pt ON pt.name = p.type_id
        WHERE {date_cond} AND {status_cond} {extra_where_sql}
        GROUP BY pt.name, pt.description
        ORDER BY sum_amount DESC
        """
        r_cat = await session.execute(text(q_cat), params)
        rows_cat = r_cat.fetchall()
        CATEGORY_RU = {"Tvorog": "Творог", "Yogurt": "Йогурт", "Tara": "Тара", "Milk": "Молоко", "Kefir": "Кефир", "Smetana": "Сметана", "Maslo": "Масло"}
        total_cat_sum = sum(float(r[2] or 0) for r in rows_cat)
        by_category = []
        for r in rows_cat:
            amt = float(r[2] or 0)
            share = round(100.0 * amt / total_cat_sum, 2) if total_cat_sum else 0
            code = (r[0] or "").strip()
            desc = (r[1] or "").strip()
            display_name = desc if desc and desc != code else (CATEGORY_RU.get(code, code) if code else "Без категории")
            by_category.append({
                "category": display_name,
                "category_code": code,
                "share_pct": share,
                "sum_amount": amt,
                "quantity": int(r[3] or 0),
            })

        v_from = _iso_to_date(df or today_str) or dt_date.today()
        v_to = _iso_to_date(dt or today_str) or dt_date.today()
        q_visits = """
        SELECT COUNT(*)::int AS total_visits,
               SUM(CASE WHEN cv.status = 'completed' THEN 1 ELSE 0 END)::int AS completed_visits
        FROM "Sales".customers_visits cv
        WHERE cv.visit_date::date >= :v_from AND cv.visit_date::date <= :v_to
        """
        r_v = await session.execute(text(q_visits), {"v_from": v_from, "v_to": v_to})
        v_row = r_v.fetchone()
        visits_total = v_row[0] or 0
        visits_completed = v_row[1] or 0
        visits_success_rate = round(100.0 * visits_completed / visits_total, 1) if visits_total else 0

        q_photos = """
        SELECT COUNT(DISTINCT cp.id)::int
        FROM "Sales".customer_photo cp
        JOIN "Sales".customers_visits cv ON cv.customer_id = cp.customer_id
        WHERE cv.visit_date::date >= :v_from AND cv.visit_date::date <= :v_to
        """
        r_ph = await session.execute(text(q_photos), {"v_from": v_from, "v_to": v_to})
        photos_count = r_ph.scalar() or 0
        photo_rate = round(100.0 * photos_count / visits_total, 1) if visits_total else 0

        r_k = await session.execute(text('SELECT COUNT(*)::int FROM "Sales".customers'))
        total_customers = r_k.scalar() or 0
        try:
            inc_r = await session.execute(text('SELECT COUNT(*)::int FROM "Sales".v_inactive_customers'))
            inactive_count = inc_r.scalar() or 0
        except Exception:
            inactive_count = 0
        active_pct = round(100.0 * (total_customers - inactive_count) / total_customers, 1) if total_customers else 0

        # --- Заказы по дням (для графика) ---
        q_orders_by_day = f"""
        SELECT o.order_date::date::text AS day,
               COUNT(*)::int AS cnt,
               COALESCE(SUM(o.total_amount), 0) AS amt
        FROM "Sales".orders o
        {cust_join} {extra_joins}
        WHERE {date_cond} AND {status_cond} {extra_where_sql}
        GROUP BY o.order_date::date
        ORDER BY day
        """
        r_obd = await session.execute(text(q_orders_by_day), params)
        orders_by_day = [{"day": str(r[0]), "count": int(r[1]), "amount": float(r[2])} for r in r_obd.fetchall()]

        # --- Заказы по статусам (для pie chart) ---
        q_orders_by_status = f"""
        SELECT o.status_code, COUNT(*)::int AS cnt, COALESCE(SUM(o.total_amount), 0) AS amt
        FROM "Sales".orders o
        {cust_join} {extra_joins}
        WHERE {date_cond} {extra_where_sql}
          AND o.status_code IS NOT NULL
        GROUP BY o.status_code
        ORDER BY cnt DESC
        """
        r_obs = await session.execute(text(q_orders_by_status), {k: v for k, v in params.items() if k != "status_codes"})
        status_labels = {"open": "Открыто", "delivery": "Доставка", "completed": "Доставлен", "canceled": "Отменён", "cancelled": "Отменён"}
        orders_by_status = [{"status": status_labels.get(str(r[0]), str(r[0])), "status_code": str(r[0]), "count": int(r[1]), "amount": float(r[2])} for r in r_obs.fetchall()]

        # --- Визиты по дням (для графика) ---
        q_visits_by_day = """
        SELECT cv.visit_date::date::text AS day,
               COUNT(*)::int AS total,
               SUM(CASE WHEN cv.status = 'completed' THEN 1 ELSE 0 END)::int AS completed
        FROM "Sales".customers_visits cv
        WHERE cv.visit_date::date >= :v_from AND cv.visit_date::date <= :v_to
        GROUP BY cv.visit_date::date
        ORDER BY day
        """
        r_vbd = await session.execute(text(q_visits_by_day), {"v_from": v_from, "v_to": v_to})
        visits_by_day = [{"day": str(r[0]), "total": int(r[1]), "completed": int(r[2])} for r in r_vbd.fetchall()]

        # --- Топ-5 клиентов по сумме ---
        q_top_clients = f"""
        SELECT c.id, COALESCE(c.name_client, c.firm_name, '') AS name,
               COUNT(DISTINCT o.order_no)::int AS orders_count,
               COALESCE(SUM(o.total_amount), 0) AS orders_sum
        FROM "Sales".orders o
        JOIN "Sales".customers c ON o.customer_id = c.id
        WHERE {date_cond} AND {status_cond}
        GROUP BY c.id, c.name_client, c.firm_name
        ORDER BY orders_sum DESC
        LIMIT 5
        """
        r_tc = await session.execute(text(q_top_clients), params)
        top_clients = [{"id": int(r[0]), "name": str(r[1]), "orders_count": int(r[2]), "orders_sum": float(r[3])} for r in r_tc.fetchall()]

        # --- Кол-во заказов и общее число ---
        q_orders_cnt = f"""
        SELECT COUNT(*)::int FROM "Sales".orders o
        {cust_join} {extra_joins}
        WHERE {date_cond} AND {status_cond} {extra_where_sql}
        """
        r_ocnt = await session.execute(text(q_orders_cnt), params)
        total_orders_count = r_ocnt.scalar() or 0

        return {
            "total_orders_sum": total_sum,
            "total_orders_count": total_orders_count,
            "by_category": by_category,
            "orders_by_day": orders_by_day,
            "orders_by_status": orders_by_status,
            "visits_by_day": visits_by_day,
            "top_clients": top_clients,
            "kpi": {
                "total_customers": total_customers,
                "visits_total": visits_total,
                "visits_completed": visits_completed,
                "visits_success_rate": visits_success_rate,
                "active_customers_pct": active_pct,
                "photo_reports_count": photos_count,
                "photo_reports_rate": photo_rate,
                "inactive_count": inactive_count,
            },
            "inactive_customers": {"count": inactive_count},
            "filters": {"date_from": df, "date_to": dt},
        }
    except Exception as e:
        return {"total_orders_sum": 0, "by_category": [], "kpi": {}, "inactive_customers": {"count": 0}, "error": str(e)[:300]}


@router.get("/dashboard/export")
async def report_dashboard_export(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    status_codes: str | None = Query(None),
    product_category: str | None = Query(None),
    territory: str | None = Query(None),
    category_client: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Экспорт сводной аналитики по категориям в Excel."""
    res = await report_dashboard(date_from, date_to, status_codes, product_category, territory, category_client, session, user)
    data = res.get("by_category") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "По категориям"
    for col, h in enumerate(["Категория", "Доля %", "Сумма", "Количество"], start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(data[:5000], start=2):
        ws.cell(row=row_idx, column=1, value=r.get("category") or "")
        ws.cell(row=row_idx, column=2, value=float(r.get("share_pct") or 0))
        ws.cell(row=row_idx, column=3, value=float(r.get("sum_amount") or 0))
        ws.cell(row=row_idx, column=4, value=int(r.get("quantity") or 0))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="dashboard_categories.xlsx"'},
    )


@router.get("/photos")
async def report_photos(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Отчёт: фотографии клиентов — табличные данные."""
    try:
        q1 = """
        SELECT COUNT(*)::int AS total_photos,
               COUNT(DISTINCT customer_id)::int AS customers_with_photos
        FROM "Sales".customer_photo
        """
        r1 = await session.execute(text(q1))
        row1 = r1.fetchone()
        total_photos = row1[0] or 0
        with_photos = row1[1] or 0
        total_cust = await session.execute(text('SELECT COUNT(*)::int FROM "Sales".customers'))
        total_c = total_cust.scalar() or 0
        without = total_c - with_photos

        # Клиенты без фото — полная таблица
        q2_sql = 'SELECT id, name_client, firm_name FROM "Sales".v_customers_without_photos'
        try:
            r2 = await session.execute(text(q2_sql))
        except Exception:
            r2 = await session.execute(text(
                'SELECT id, name_client, firm_name FROM "Sales".customers c '
                'WHERE NOT EXISTS (SELECT 1 FROM "Sales".customer_photo cp WHERE cp.customer_id = c.id) '
                'ORDER BY c.name_client'
            ))
        rows2 = r2.fetchall()
        without_list = [{"id": r[0], "name_client": r[1] or "", "firm_name": r[2] or "", "name": (r[1] or r[2] or "")} for r in rows2]

        # Все фото — полная таблица с клиентом, датой, автором, описанием, ссылкой
        q3 = """
        SELECT cp.id, cp.customer_id, COALESCE(c.name_client, c.firm_name, '') AS customer_name,
               cp.photo_path, cp.original_filename, cp.file_size,
               cp.description, cp.uploaded_by, cp.uploaded_at,
               u.fio AS uploader_fio
        FROM "Sales".customer_photo cp
        JOIN "Sales".customers c ON c.id = cp.customer_id
        LEFT JOIN "Sales".users u ON cp.uploaded_by = u.login
        ORDER BY cp.uploaded_at DESC NULLS LAST
        """
        r3 = await session.execute(text(q3))
        all_photos = []
        for row in r3.fetchall():
            all_photos.append({
                "id": row[0],
                "customer_id": row[1],
                "customer_name": str(row[2] or ""),
                "photo_path": row[3] or "",
                "original_filename": row[4] or "",
                "file_size": row[5] or 0,
                "description": row[6] or "",
                "uploaded_by": row[7] or "",
                "uploader_fio": str(row[8] or row[7] or ""),
                "uploaded_at": row[8].isoformat() if hasattr(row[8], "isoformat") else str(row[8] or ""),
                "uploader_fio": str(row[9] or row[7] or ""),
            })

        # Фото по клиентам — сводка
        q4 = """
        SELECT c.id, COALESCE(c.name_client, c.firm_name, '') AS name,
               COUNT(cp.id)::int AS photo_count,
               MAX(cp.uploaded_at) AS last_upload
        FROM "Sales".customers c
        LEFT JOIN "Sales".customer_photo cp ON cp.customer_id = c.id
        GROUP BY c.id, c.name_client, c.firm_name
        ORDER BY photo_count DESC, name
        """
        r4 = await session.execute(text(q4))
        by_customer = []
        for row in r4.fetchall():
            by_customer.append({
                "id": row[0],
                "name": str(row[1] or ""),
                "photo_count": int(row[2] or 0),
                "last_upload": row[3].isoformat() if hasattr(row[3], "isoformat") else (str(row[3]) if row[3] else ""),
            })

        return {
            "statistics": {
                "total_photos": total_photos,
                "customers_with_photos": with_photos,
                "customers_without_photos": without,
                "total_customers": total_c,
            },
            "all_photos": all_photos,
            "by_customer": by_customer,
            "customers_without_photos": without_list,
        }
    except Exception as e:
        return {"statistics": {}, "all_photos": [], "by_customer": [], "customers_without_photos": [], "error": str(e)[:200]}


@router.get("/photos/export")
async def report_photos_export(
    session: AsyncSession = Depends(get_db_session),
    user: User = Depends(get_current_user),
):
    """Экспорт отчёта по фотографиям клиентов в Excel."""
    res = await report_photos(session, user)
    stats = res.get("statistics") or {}
    without = res.get("customers_without_photos") or []
    recent = res.get("recent_uploads") or []
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Статистика"
    ws1.cell(row=1, column=1, value="Всего фото")
    ws1.cell(row=1, column=2, value=stats.get("total_photos") or 0)
    ws1.cell(row=2, column=1, value="Клиентов с фото")
    ws1.cell(row=2, column=2, value=stats.get("customers_with_photos") or 0)
    ws1.cell(row=3, column=1, value="Без фото")
    ws1.cell(row=3, column=2, value=stats.get("customers_without_photos") or 0)
    ws2 = wb.create_sheet("Клиенты без фото")
    ws2.cell(row=1, column=1, value="Клиент")
    for row_idx, c in enumerate(without[:5000], start=2):
        ws2.cell(row=row_idx, column=1, value=(c.get("name_client") or c.get("firm_name") or c.get("name") or ""))
    ws3 = wb.create_sheet("Последние загрузки")
    for col, h in enumerate(["Клиент", "Дата загрузки", "Загружено"], start=1):
        ws3.cell(row=1, column=col, value=h)
    for row_idx, p in enumerate(recent[:500], start=2):
        ws3.cell(row=row_idx, column=1, value=p.get("customer_name") or "")
        ws3.cell(row=row_idx, column=2, value=p.get("uploaded_at") or "")
        ws3.cell(row=row_idx, column=3, value=p.get("uploaded_by") or "")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_photos.xlsx"'},
    )

"""
Визиты (visits): поиск, календарь, CRUD. Модель CustomerVisit.
"""
import io
from datetime import date, datetime, time
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook

from src.database.connection import get_db_session
from src.database.models import CustomerVisit, Customer, User
from src.core.deps import get_current_user
from src.database.models import User as UserModel

router = APIRouter()

VISITS_EXPORT_HEADERS = ["Дата", "Время", "Клиент", "Статус", "Ответственный", "Комментарий"]
STATUS_RU = {"planned": "Запланирован", "completed": "Завершён", "cancelled": "Отменён", "postponed": "На рассмотрении"}


class VisitCreate(BaseModel):
    customer_id: int
    visit_date: str  # YYYY-MM-DD
    visit_time: str | None = None
    status: str = "planned"
    responsible_login: str | None = None
    comment: str | None = None


class VisitUpdate(BaseModel):
    visit_date: str | None = None
    visit_time: str | None = None
    status: str | None = None
    responsible_login: str | None = None
    comment: str | None = None


@router.post("/visits")
async def create_visit(
    body: VisitCreate,
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Создать новый визит."""
    # Проверить что клиент существует
    r = await session.execute(select(Customer).where(Customer.id == body.customer_id))
    customer = r.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Клиент не найден")

    # Создать визит
    visit = CustomerVisit(
        customer_id=body.customer_id,
        visit_date=date.fromisoformat(body.visit_date[:10]),
        visit_time=None,
        status=body.status or "planned",
        responsible_login=body.responsible_login or user.login,
        comment=body.comment,
        created_by=user.login,
        updated_by=user.login,
    )

    # Обработать время визита
    if body.visit_time and body.visit_time.strip():
        parts = body.visit_time.strip()[:5].split(":")
        if len(parts) == 2:
            try:
                visit.visit_time = time(int(parts[0]), int(parts[1]))
            except (ValueError, TypeError):
                pass

    session.add(visit)
    await session.commit()
    await session.refresh(visit)
    return {"id": visit.id, "message": "created"}


@router.get("/visits/search")
async def search_visits(
    customer_id: int | None = Query(None),
    customer_name: str | None = Query(None),
    status: str | None = Query(None, description="Один или несколько статусов через запятую (planned,completed,cancelled,postponed)"),
    responsible_login: str | None = Query(None),
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    limit: int = Query(50, le=200),
    offset: int = Query(0, ge=0),
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Поиск визитов с фильтрами."""
    q = (
        select(CustomerVisit, Customer.name_client, Customer.firm_name, User.fio)
        .join(Customer, CustomerVisit.customer_id == Customer.id)
        .outerjoin(User, CustomerVisit.responsible_login == User.login)
        .order_by(CustomerVisit.visit_date.desc(), CustomerVisit.visit_time.desc().nullslast())
    )
    if customer_id is not None:
        q = q.where(CustomerVisit.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        q = q.where(
            (Customer.name_client.ilike(f"%{name}%")) | (Customer.firm_name.ilike(f"%{name}%"))
        )
    if status and status.strip():
        parts = [s.strip() for s in status.split(",") if s.strip()]
        if parts:
            q = q.where(CustomerVisit.status.in_(parts))
    if responsible_login and responsible_login.strip():
        q = q.where(CustomerVisit.responsible_login == responsible_login.strip())
    if from_date:
        try:
            d = date.fromisoformat(from_date[:10])
            q = q.where(CustomerVisit.visit_date >= d)
        except (ValueError, TypeError):
            pass
    if to_date:
        try:
            d = date.fromisoformat(to_date[:10])
            q = q.where(CustomerVisit.visit_date <= d)
        except (ValueError, TypeError):
            pass
    count_q = select(func.count()).select_from(CustomerVisit)
    if customer_id is not None:
        count_q = count_q.where(CustomerVisit.customer_id == customer_id)
    if customer_name and customer_name.strip():
        count_q = count_q.join(Customer, CustomerVisit.customer_id == Customer.id)
        name = customer_name.strip()
        count_q = count_q.where(
            (Customer.name_client.ilike(f"%{name}%")) | (Customer.firm_name.ilike(f"%{name}%"))
        )
    if status and status.strip():
        parts = [s.strip() for s in status.split(",") if s.strip()]
        if parts:
            count_q = count_q.where(CustomerVisit.status.in_(parts))
    if responsible_login and responsible_login.strip():
        count_q = count_q.where(CustomerVisit.responsible_login == responsible_login.strip())
    if from_date:
        try:
            count_q = count_q.where(CustomerVisit.visit_date >= date.fromisoformat(from_date[:10]))
        except (ValueError, TypeError):
            pass
    if to_date:
        try:
            count_q = count_q.where(CustomerVisit.visit_date <= date.fromisoformat(to_date[:10]))
        except (ValueError, TypeError):
            pass
    total = (await session.execute(count_q)).scalar() or 0
    q = q.offset(offset).limit(limit)
    result = await session.execute(q)
    rows = result.all()
    data = []
    for v, name_client, firm_name, resp_fio in rows:
        data.append({
            "id": v.id,
            "customer_id": v.customer_id,
            "customer_name": (name_client or firm_name or "").strip() or f"Клиент #{v.customer_id}",
            "visit_date": v.visit_date.isoformat() if v.visit_date else None,
            "visit_time": v.visit_time.strftime("%H:%M") if v.visit_time else None,
            "status": v.status,
            "responsible_login": v.responsible_login,
            "responsible_name": resp_fio or v.responsible_login or "",
            "comment": v.comment,
        })
    return {"total": total, "data": data}


@router.get("/visits/export")
async def export_visits_excel(
    customer_id: int | None = Query(None),
    customer_name: str | None = Query(None),
    status: str | None = Query(None),
    responsible_login: str | None = Query(None),
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Выгрузка визитов в Excel по тем же фильтрам, что и поиск."""
    q = (
        select(CustomerVisit, Customer.name_client, Customer.firm_name, User.fio)
        .join(Customer, CustomerVisit.customer_id == Customer.id)
        .outerjoin(User, CustomerVisit.responsible_login == User.login)
        .order_by(CustomerVisit.visit_date.desc(), CustomerVisit.visit_time.desc().nullslast())
    )
    if customer_id is not None:
        q = q.where(CustomerVisit.customer_id == customer_id)
    if customer_name and customer_name.strip():
        name = customer_name.strip()
        q = q.where((Customer.name_client.ilike(f"%{name}%")) | (Customer.firm_name.ilike(f"%{name}%")))
    if status and status.strip():
        parts = [s.strip() for s in status.split(",") if s.strip()]
        if parts:
            q = q.where(CustomerVisit.status.in_(parts))
    if responsible_login and responsible_login.strip():
        q = q.where(CustomerVisit.responsible_login == responsible_login.strip())
    if from_date:
        try:
            q = q.where(CustomerVisit.visit_date >= date.fromisoformat(from_date[:10]))
        except (ValueError, TypeError):
            pass
    if to_date:
        try:
            q = q.where(CustomerVisit.visit_date <= date.fromisoformat(to_date[:10]))
        except (ValueError, TypeError):
            pass
    result = await session.execute(q.limit(50000))
    rows = result.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Визиты"
    for col, name in enumerate(VISITS_EXPORT_HEADERS, start=1):
        ws.cell(row=1, column=col, value=name)
    for row_idx, (v, name_client, firm_name, resp_fio) in enumerate(rows, start=2):
        customer_name_val = (name_client or firm_name or "").strip() or (f"Клиент #{v.customer_id}" if v.customer_id else "")
        status_ru = STATUS_RU.get(v.status, v.status or "")
        row_data = [
            v.visit_date.isoformat() if v.visit_date else "",
            v.visit_time.strftime("%H:%M") if v.visit_time else "",
            customer_name_val,
            status_ru,
            resp_fio or v.responsible_login or "",
            v.comment or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="visits.xlsx"'},
    )


@router.get("/visits/calendar")
async def visits_calendar(
    year: int = Query(..., ge=2020, le=2030),
    month: int = Query(..., ge=1, le=12),
    responsible_login: str | None = Query(None),
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Календарь визитов по месяцу."""
    from_date = date(year, month, 1)
    if month == 12:
        to_date = date(year + 1, 1, 1)
    else:
        to_date = date(year, month + 1, 1)
    q = (
        select(CustomerVisit, Customer.name_client, Customer.firm_name, User.fio)
        .join(Customer, CustomerVisit.customer_id == Customer.id)
        .outerjoin(User, CustomerVisit.responsible_login == User.login)
        .where(CustomerVisit.visit_date >= from_date, CustomerVisit.visit_date < to_date)
        .order_by(CustomerVisit.visit_date, CustomerVisit.visit_time)
    )
    if responsible_login and responsible_login.strip():
        q = q.where(CustomerVisit.responsible_login == responsible_login.strip())
    result = await session.execute(q)
    rows = result.all()
    events = []
    for v, name_client, firm_name, resp_fio in rows:
        d = v.visit_date.isoformat() if v.visit_date else ""
        t = v.visit_time.strftime("%H:%M") if v.visit_time else ""
        customer_name = (name_client or firm_name or "").strip() or f"Клиент #{v.customer_id}"
        events.append({
            "id": v.id,
            "date": d,
            "time": t,
            "customer_id": v.customer_id,
            "customer_name": customer_name,
            "status": v.status,
            "responsible_login": v.responsible_login,
            "responsible_name": resp_fio or v.responsible_login or "",
        })
    return {"events": events}


@router.get("/visits/{visit_id}")
async def get_visit(
    visit_id: int,
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Получить визит по ID."""
    r = await session.execute(
        select(CustomerVisit, Customer.name_client, Customer.firm_name, User.fio)
        .join(Customer, CustomerVisit.customer_id == Customer.id)
        .outerjoin(User, CustomerVisit.responsible_login == User.login)
        .where(CustomerVisit.id == visit_id)
    )
    row = r.one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Визит не найден")
    v, name_client, firm_name, resp_fio = row
    return {
        "id": v.id,
        "customer_id": v.customer_id,
        "customer_name": (name_client or firm_name or "").strip() or f"Клиент #{v.customer_id}",
        "visit_date": v.visit_date.isoformat() if v.visit_date else None,
        "visit_time": v.visit_time.strftime("%H:%M") if v.visit_time else None,
        "status": v.status,
        "responsible_login": v.responsible_login,
        "responsible_name": resp_fio or v.responsible_login or "",
        "comment": v.comment,
    }


@router.put("/visits/{visit_id}")
async def update_visit(
    visit_id: int,
    body: VisitUpdate,
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Обновить визит."""
    r = await session.execute(select(CustomerVisit).where(CustomerVisit.id == visit_id))
    v = r.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Визит не найден")
    if body.visit_date is not None:
        v.visit_date = date.fromisoformat(body.visit_date[:10])
    if body.visit_time is not None:
        if not body.visit_time or not body.visit_time.strip():
            v.visit_time = None
        else:
            parts = body.visit_time.strip()[:5].split(":")
            if len(parts) == 2:
                try:
                    v.visit_time = time(int(parts[0]), int(parts[1]))
                except (ValueError, TypeError):
                    pass
    if body.status is not None:
        v.status = body.status
    if body.responsible_login is not None:
        v.responsible_login = body.responsible_login or None
    if body.comment is not None:
        v.comment = body.comment or None
    v.updated_by = user.login
    await session.commit()
    await session.refresh(v)
    return {"id": v.id, "message": "updated"}


@router.delete("/visits/{visit_id}")
async def delete_visit(
    visit_id: int,
    session: AsyncSession = Depends(get_db_session),
    user: UserModel = Depends(get_current_user),
):
    """Удалить визит."""
    r = await session.execute(select(CustomerVisit).where(CustomerVisit.id == visit_id))
    v = r.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Визит не найден")
    await session.delete(v)
    await session.commit()
    return {"message": "deleted"}
